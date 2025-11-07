"""
Módulo de gerenciamento de banco de dados para conciliação contábil.
Responsável por importar, processar e exportar dados de diferentes fontes
para realizar a conciliação entre sistemas financeiros e contábeis.
"""

import sqlite3
from pathlib import Path
from config.settings import Settings
from config.logger import configure_logger
from .exceptions import (
    PlanilhaFormatacaoErradaError,
    InvalidDataFormat,
    ResultsSaveError,
    ExcecaoNaoMapeadaError
)
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from difflib import get_close_matches
from workalendar.america import Brazil
from datetime import datetime, timedelta
import calendar
import locale
import pandas as pd
import xml.etree.ElementTree as ET
import numpy as np
import openpyxl
import re

# Configura o logger para registrar eventos
logger = configure_logger()
locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
class DatabaseManager:
    """
    Gerenciador de banco de dados para conciliação contábil.
    Implementa padrão Singleton para garantir apenas uma instância.
    """
    
    # Implementação do padrão Singleton
    _instance = None
    
    def __new__(cls):
        """
        Implementa o padrão Singleton para garantir apenas uma instância.
        
        Returns:
            DatabaseManager: Instância única da classe
        """
        if cls._instance is None:
            cls._instance = super(DatabaseManager, cls).__new__(cls)
            cls._instance._initialized = False
        return cls._instance
    
    def __init__(self):
        """
        Inicializa o gerenciador de banco de dados.
        Evita múltiplas inicializações no padrão Singleton.
        """
        # Verifica se já foi inicializado
        if self._initialized:
            return
            
        self.settings = Settings()  # Carrega configurações
        self.conn = None  # Conexão com o banco
        self.logger = configure_logger()  # Logger específico da classe
        self._initialize_database()  # Inicializa o banco de dados
        self._initialized = True  # Marca como inicializado

    def _initialize_database(self):
        """
        Inicializa o banco de dados SQLite e cria as tabelas necessárias.
        
        Raises:
            Exception: Se ocorrer erro na inicialização do banco
        """
        try:
            # Conecta ao banco SQLite
            self.conn = sqlite3.connect(self.settings.DB_PATH, timeout=10)
            cursor = self.conn.cursor()
            
            # Cria tabela financeiro se não existir
            cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS {self.settings.TABLE_FINANCEIRO} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    fornecedor TEXT,
                    titulo TEXT,
                    parcela TEXT,
                    tipo_titulo TEXT,
                    data_emissao TEXT DEFAULT NULL,
                    data_vencimento TEXT DEFAULT NULL,
                    valor_original REAL DEFAULT 0,
                    tit_vencidos_valor_nominal REAL DEFAULT 0,  -- NOVA COLUNA
                    titulos_a_vencer_valor_nominal REAL DEFAULT 0,  -- NOVA COLUNA
                    situacao TEXT,
                    conta_contabil TEXT,
                    centro_custo TEXT,
                    excluido BOOLEAN DEFAULT 0,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Cria tabela modelo1 (contábil) se não existir
            cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS {self.settings.TABLE_MODELO1} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    conta_contabil TEXT,
                    descricao_conta TEXT,
                    codigo_fornecedor TEXT,
                    descricao_fornecedor TEXT,
                    saldo_anterior REAL DEFAULT 0,
                    debito REAL DEFAULT 0,
                    credito REAL DEFAULT 0,
                    saldo_atual REAL DEFAULT 0,
                    tipo_fornecedor TEXT,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Cria tabela contas_itens se não existir
            cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS {self.settings.TABLE_CONTAS_ITENS} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    conta_contabil TEXT,
                    descricao_item TEXT,
                    codigo_fornecedor TEXT,
                    descricao_fornecedor TEXT,
                    saldo_anterior REAL DEFAULT 0,
                    debito REAL DEFAULT 0,
                    credito REAL DEFAULT 0,
                    saldo_atual REAL DEFAULT 0,
                    item TEXT DEFAULT '',
                    quantidade REAL DEFAULT 1,
                    valor_unitario REAL DEFAULT 0,
                    valor_total REAL DEFAULT 0,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            # Cria tabela adiantamento se não existir
            cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS {self.settings.TABLE_ADIANTAMENTO} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    conta_contabil TEXT,
                    descricao_item TEXT,
                    codigo_fornecedor TEXT,
                    descricao_fornecedor TEXT,
                    saldo_anterior REAL DEFAULT 0,
                    debito REAL DEFAULT 0,
                    credito REAL DEFAULT 0,
                    saldo_atual REAL DEFAULT 0,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Cria tabela Adiantamento financeiro se não existir
            cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS {self.settings.TABLE_RESULTADO_ADIANTAMENTO} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    codigo_fornecedor TEXT,
                    descricao_fornecedor TEXT,
                    total_financeiro REAL DEFAULT 0,
                    total_contabil REAL DEFAULT 0,
                    diferenca REAL DEFAULT 0,
                    status TEXT CHECK(status IN ('Conferido', 'Divergente', 'Pendente')),
                    detalhes TEXT,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Cria tabela resultado (concatenação) se não existir
            cursor.execute(f"""
                CREATE TABLE IF NOT EXISTS {self.settings.TABLE_RESULTADO} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    codigo_fornecedor TEXT,
                    descricao_fornecedor TEXT,
                    saldo_contabil REAL DEFAULT 0,
                    saldo_financeiro REAL DEFAULT 0,
                    diferenca REAL DEFAULT 0,
                    status TEXT CHECK(status IN ('Conferido', 'Divergente', 'Pendente')),
                    detalhes TEXT,
                    ordem_importancia INTEGER,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            
            
            # Função auxiliar para garantir que colunas existam nas tabelas
            def ensure_column(table, column, type_):
                """Garante que uma coluna exista na tabela especificada."""
                cursor.execute(f"PRAGMA table_info({table})")
                cols = [c[1] for c in cursor.fetchall()]
                if column not in cols:
                    cursor.execute(f"ALTER TABLE {table} ADD COLUMN {column} {type_}")
            
            # Garante que colunas importantes existam
            ensure_column(self.settings.TABLE_FINANCEIRO, 'tit_vencidos_valor_nominal', 'REAL')
            ensure_column(self.settings.TABLE_FINANCEIRO, 'titulos_a_vencer_valor_nominal', 'REAL')
            ensure_column(self.settings.TABLE_CONTAS_ITENS, 'codigo_fornecedor', 'TEXT')
            ensure_column(self.settings.TABLE_CONTAS_ITENS, 'descricao_fornecedor', 'TEXT')
            ensure_column(self.settings.TABLE_MODELO1, 'codigo_fornecedor', 'TEXT')
            ensure_column(self.settings.TABLE_MODELO1, 'descricao_fornecedor', 'TEXT')
            ensure_column(self.settings.TABLE_RESULTADO, 'ordem_importancia', 'INTEGER')
            
            self.conn.commit()  # Confirma as alterações
            logger.info("Banco de dados inicializado com sucesso")
            
        except Exception as e:
            error_msg = f"Erro ao inicializar banco de dados: {e}"
            logger.error(error_msg)
            raise ExcecaoNaoMapeadaError(error_msg) from e

    def aplicar_sugestoes_colunas(self, df, missing_mappings):
        """
        Aplica sugestões automáticas para mapeamento de colunas faltantes.
        
        Args:
            df: DataFrame com os dados
            missing_mappings: Lista de colunas que faltam mapeamento
            
        Returns:
            DataFrame: DataFrame com colunas renomeadas conforme sugestões
        """
        try:
            candidates = df.columns.tolist()
            lower_map = {c.lower(): c for c in candidates}  # Mapa case-insensitive

            # Mapeamento manual pré-definido para colunas comuns
            manual_mapping = {
                'Codigo-Nome do Fornecedor': 'fornecedor',
                'Prf-Numero Parcela': 'titulo', 
                'Tp': 'tipo_titulo',
                'Data de Emissao': 'data_emissao',
                'Data Emissão': 'data_emissao',
                'Data de Vencto': 'data_vencimento',
                'Data Vencimento': 'data_vencimento',
                'Valor Original': 'valor_original',
                'Tit Vencidos Valor nominal': 'saldo_devedor',
                'Natureza': 'situacao',
                'Porta- dor': 'centro_custo',
                'Codigo.1': 'codigo_fornecedor',
                'Descricao.1': 'descricao_fornecedor',
                'Conta': 'conta_contabil',
                'Descricao': 'descricao_conta',
                'Descricao': 'descricao_item',  # Para contas_itens
                'Codigo': 'conta_contabil'  # Para contas_itens e adiantamento
            }

            # Aplica mapeamento manual primeiro
            for src, dest in manual_mapping.items():
                if src in df.columns and dest in missing_mappings:
                    df.rename(columns={src: dest}, inplace=True)
                    logger.warning(f"Mapeamento manual aplicado: '{src}' → '{dest}'")
                    if dest in missing_mappings:
                        missing_mappings.remove(dest)

            # Tenta encontrar correspondências automáticas para colunas restantes
            for db_col in list(missing_mappings):
                # Busca por correspondência case-insensitive
                if db_col.lower() in lower_map:
                    match = lower_map[db_col.lower()]
                    logger.warning(f"Sugestão aplicada: '{match}' → '{db_col}' (case-insensitive match)")
                    df.rename(columns={match: db_col}, inplace=True)
                    if db_col in missing_mappings:
                        missing_mappings.remove(db_col)
                    continue

                # Busca por correspondência fuzzy (similaridade)
                similar = get_close_matches(db_col, candidates, n=1, cutoff=0.6)
                if similar:
                    match = similar[0]
                    logger.warning(f"Sugestão aplicada: '{match}' → '{db_col}'")
                    df.rename(columns={match: db_col}, inplace=True)
                    if db_col in missing_mappings:
                        missing_mappings.remove(db_col)
                    continue

                # Busca por correspondência fuzzy case-insensitive
                similar_lower = get_close_matches(db_col.lower(), list(lower_map.keys()), n=1, cutoff=0.6)
                if similar_lower:
                    match = lower_map[similar_lower[0]]
                    logger.warning(f"Sugestão aplicada: '{match}' → '{db_col}' (fuzzy case-insensitive)")
                    df.rename(columns={match: db_col}, inplace=True)
                    if db_col in missing_mappings:
                        missing_mappings.remove(db_col)

            return df
        except Exception as e:
            error_msg = f"Erro ao aplicar sugestões de colunas: {e}"
            logger.error(error_msg)
            raise PlanilhaFormatacaoErradaError(error_msg) from e

    def import_from_excel(self, file_path, table_name):
        """
        Importa dados de arquivo Excel/TXT/XML para a tabela especificada.
        """
        try:
            filename = Path(file_path).stem.lower()

            # Determina a tabela destino baseada no nome do arquivo
            if 'ctbr100' in filename:
                table_name = self.settings.TABLE_ADIANTAMENTO
            elif 'ctbr140' in filename:
                table_name = self.settings.TABLE_CONTAS_ITENS
            elif 'ctbr040' in filename:
                table_name = self.settings.TABLE_MODELO1
            elif 'finr150' in filename:
                table_name = self.settings.TABLE_FINANCEIRO

            ext = Path(file_path).suffix.lower()
            
            # Lê o arquivo conforme o formato
            if ext == ".xlsx":
                # Lê as primeiras linhas para diagnóstico
                df_sample = pd.read_excel(file_path, nrows=5)
                logger.info(f"Primeiras 5 linhas do arquivo {file_path}:")
                logger.info(df_sample.to_string())
                
                # Lê o arquivo completo a partir da linha 2 (header=1)
                df = pd.read_excel(file_path, header=1)

            elif ext == ".xml":
                try:
                    df = DatabaseManager.read_spreadsheetml(file_path)
                except Exception as e:
                    error_msg = f"Falha ao ler {file_path} como SpreadsheetML: {e}"
                    logger.error(error_msg)
                    raise InvalidDataFormat(error_msg, tipo_dado="XML") from e

            elif ext == ".txt":
                try:
                    df = pd.read_csv(file_path, sep=";", encoding="latin1", header=1)
                except Exception:
                    df = pd.read_csv(file_path, sep="\t", encoding="latin1", header=1)

            else:
                error_msg = f"Formato de arquivo não suportado: {ext}"
                logger.error(error_msg)
                raise InvalidDataFormat(error_msg, tipo_dado=ext)

            logger.info(f"Colunas originais em {file_path}: {df.columns.tolist()}")
            logger.info(f"Primeiras linhas dos dados:")
            logger.info(df.head().to_string())

            # Limpa caracteres especiais dos nomes das colunas
            df.columns = df.columns.str.replace(r'_x000D_\n', ' ', regex=True).str.strip()
            logger.info(f"Colunas após limpeza: {df.columns.tolist()}")

            # Aplica mapeamento de colunas
            column_mapping = self._get_column_mapping(Path(file_path))
            
            # Verifica se column_mapping é um dicionário válido
            if not isinstance(column_mapping, dict):
                logger.warning(f"Mapeamento de colunas inválido para {file_path}, usando mapeamento vazio")
                column_mapping = {}
                
            if column_mapping:  # Só aplica se houver mapeamento
                df.rename(columns=column_mapping, inplace=True)
                
            logger.info(f"Colunas após mapeamento: {df.columns.tolist()}")
            logger.info(f"Amostra dos dados após mapeamento:")
            logger.info(df.head().to_string())

            # Verifica colunas obrigatórias
            expected_columns = self.get_expected_columns(table_name)
            missing_mappings = [col for col in expected_columns if col not in df.columns]

            if missing_mappings:
                logger.warning(f"Colunas mapeadas não encontradas: {missing_mappings}")
                df = self.aplicar_sugestoes_colunas(df, missing_mappings)
                remaining_missing = [col for col in expected_columns if col not in df.columns]

                if remaining_missing:
                    #  TRECHO CORRIGIDO: Tratamento específico para colunas ausentes
                    # Tenta criar colunas ausentes com valores padrão
                    if 'parcela' in remaining_missing and 'titulo' in df.columns:
                        df['parcela'] = df['titulo'].astype(str).str.extract(r'(\d+)$').fillna('1')
                        logger.warning("Coluna 'parcela' criada a partir do título")
                        remaining_missing.remove('parcela')

                    if 'conta_contabil' in remaining_missing:
                        df['conta_contabil'] = 'CONTA_NAO_IDENTIFICADA'
                        logger.warning("Coluna 'conta_contabil' preenchida com valor padrão para arquivo financeiro")
                        remaining_missing.remove('conta_contabil')

                    #  NOVO: Para saldo_devedor, calculamos a partir das novas colunas J e K
                    if 'saldo_devedor' in remaining_missing:
                        # Verifica se temos as colunas J e K para calcular o saldo_devedor
                        if 'tit_vencidos_valor_nominal' in df.columns and 'titulos_a_vencer_valor_nominal' in df.columns:
                            df['saldo_devedor'] = df['tit_vencidos_valor_nominal'].fillna(0) + df['titulos_a_vencer_valor_nominal'].fillna(0)
                            logger.warning("Coluna 'saldo_devedor' calculada a partir de tit_vencidos_valor_nominal + titulos_a_vencer_valor_nominal")
                            remaining_missing.remove('saldo_devedor')
                        else:
                            # Se não temos J e K, usa valor_original como fallback
                            df['saldo_devedor'] = df.get('valor_original', 0)
                            logger.warning("Coluna 'saldo_devedor' preenchida com valor_original como fallback")
                            remaining_missing.remove('saldo_devedor')

                    if remaining_missing:
                        error_msg = f"Colunas obrigatórias ausentes após tratamento: {remaining_missing}"
                        logger.error(error_msg)
                        raise PlanilhaFormatacaoErradaError(error_msg, caminho_arquivo=file_path)

            # Limpa e prepara os dados - com diagnóstico detalhado para datas
            logger.info("Iniciando limpeza dos dados...")
            df = self._clean_dataframe(df, table_name.lower())
            
            # Verificação específica das colunas de data
            for date_col in ['data_emissao', 'data_vencimento']:
                if date_col in df.columns:
                    logger.info(f"Coluna {date_col} - Valores únicos: {df[date_col].unique()}")
                    logger.info(f"Coluna {date_col} - Tipos: {df[date_col].dtype}")
                    logger.info(f"Coluna {date_col} - Não nulos: {df[date_col].notna().sum()}")

            # Mantém apenas colunas que existem na tabela destino
            table_columns = [col[1] for col in self.conn.execute(f"PRAGMA table_info({table_name})").fetchall()]
            keep = [col for col in df.columns if col in table_columns]
            df = df[keep]

            for col in table_columns:
                if col not in df.columns:
                    if col == 'excluido':
                        df[col] = 0  # Valor padrão para coluna excluido
                    else:
                        df[col] = None  # Valor padrão para outras colunas

            # Insere dados no banco
            df.to_sql(table_name, self.conn, if_exists='replace', index=False)
            logger.info(f"Dados importados para '{table_name}' com sucesso.")
            return True

        except (PlanilhaFormatacaoErradaError, InvalidDataFormat) as e:
            logger.error(f"Falha ao importar {file_path}: {e}")
            return False
        except Exception as e:
            error_msg = f"Erro inesperado ao importar {file_path}: {e}"
            logger.error(error_msg, exc_info=True)
            raise ExcecaoNaoMapeadaError(error_msg) from e

    @staticmethod
    def read_spreadsheetml(path: str) -> pd.DataFrame:
        """
        Lê arquivos XML no formato SpreadsheetML.
        
        Args:
            path: Caminho do arquivo XML
            
        Returns:
            DataFrame: Dados lidos do arquivo XML
            
        Raises:
            ValueError: Se não encontrar cabeçalho e dados suficientes
        """
        try:
            ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
            tree = ET.parse(path)
            root = tree.getroot()

            rows = []
            for row in root.findall(".//ss:Row", ns):
                values = []
                for cell in row.findall("ss:Cell", ns):
                    data = cell.find("ss:Data", ns)
                    values.append(data.text if data is not None else None)
                rows.append(values)

            if not rows or len(rows) < 2:
                raise ValueError("Não foi possível encontrar cabeçalho e dados no arquivo XML")

            # Pula a primeira linha (título "Item Conta")
            header = rows[1]
            data = rows[2:]

            df = pd.DataFrame(data, columns=header)

            # Deduplicar nomes de colunas manualmente
            counts = {}
            new_columns = []
            for col in df.columns:
                if col not in counts:
                    counts[col] = 0
                    new_columns.append(col)
                else:
                    counts[col] += 1
                    new_columns.append(f"{col}_{counts[col]}")

            df.columns = new_columns
            return df
        except Exception as e:
            error_msg = f"Erro ao ler arquivo SpreadsheetML {path}: {e}"
            logger.error(error_msg)
            raise InvalidDataFormat(error_msg, tipo_dado="XML") from e

    def get_expected_columns(self, table_name):
        """
        Retorna lista de colunas esperadas para cada tipo de tabela.
        """
        if table_name == self.settings.TABLE_FINANCEIRO:
            return [
                'fornecedor', 'titulo', 'parcela', 'tipo_titulo',
                'data_emissao', 'data_vencimento', 'valor_original',
                'tit_vencidos_valor_nominal', 'titulos_a_vencer_valor_nominal',  # NOVAS COLUNAS
                'situacao', 'conta_contabil', 'centro_custo'
            ]
        elif table_name == self.settings.TABLE_MODELO1:
            return [
                'conta_contabil', 'descricao_conta',
                'saldo_anterior', 'debito', 'credito', 'saldo_atual'
            ]
        elif table_name == self.settings.TABLE_CONTAS_ITENS:
            return [
                'conta_contabil', 'descricao_item',
                'codigo_fornecedor', 'descricao_fornecedor',
                'saldo_anterior', 'debito', 'credito', 'saldo_atual'
            ]
        elif table_name == self.settings.TABLE_ADIANTAMENTO:
            return [
                'conta_contabil', 'descricao_item',
                'codigo_fornecedor', 'descricao_fornecedor',
                'saldo_anterior', 'debito', 'credito', 'saldo_atual'
            ]
        else:
            error_msg = f"Tabela desconhecida: {table_name}"
        logger.error(error_msg)
        raise ValueError(error_msg)
    
    def _clean_dataframe(self, df, sheet_type):
        """
        Executa limpeza geral do DataFrame baseado no tipo de planilha.
        
        Args:
            df: DataFrame a ser limpo
            sheet_type: Tipo de planilha ('financeiro', 'modelo1', 'contas_itens', 'adiantamento')
            
        Returns:
            DataFrame: DataFrame limpo
            
        Raises:
            Exception: Se ocorrer erro na limpeza
        """
        try:
            # Limpa strings e remove valores vazios
            df = df.map(lambda x: str(x).strip() if pd.notna(x) else x)
            df = df.replace(['nan', 'None', ''], np.nan)
            df = df.dropna(how='all')  # Remove linhas completamente vazias
            
            # Aplica limpeza específica por tipo de planilha
            if sheet_type == 'financeiro':
                df = self._clean_financeiro_data(df)
            elif sheet_type == 'modelo1':
                df = self._clean_modelo1_data(df)
            elif sheet_type == 'contas_itens':
                df = self._clean_contas_itens_data(df)
            elif sheet_type == 'adiantamento':  
                df = self._clean_adiantamento_data(df)
            
            df = df.drop_duplicates()  # Remove duplicatas
            logger.info(f"DataFrame limpo - shape final: {df.shape}")
            return df
        except Exception as e:
            error_msg = f"Erro na limpeza dos dados ({sheet_type}): {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise InvalidDataFormat(error_msg, tipo_dado=sheet_type) from e

    def _clean_financeiro_data(self, df):
        """
        Limpeza específica para dados financeiros.
        Aplica separação de código e descrição durante a importação.
        """
        try:
            # Diagnóstico inicial das colunas de data
            logger.info("Iniciando limpeza de dados financeiros...")
            for date_col in ['data_emissao', 'data_vencimento']:
                if date_col in df.columns:
                    logger.info(f"Coluna {date_col} antes da limpeza:")
                    logger.info(f"  Tipo: {df[date_col].dtype}")
                    logger.info(f"  Primeiros valores: {df[date_col].head().tolist()}")
                    logger.info(f"  Valores únicos: {df[date_col].unique()[:5]}")

            # APLICA SEPARAÇÃO DE CÓDIGO E DESCRIÇÃO DURANTE A IMPORTÇÃO
            if 'fornecedor' in df.columns:
                logger.info("Aplicando separação de código e descrição na coluna 'fornecedor'")
                df = self.separar_codigo_descricao(df, 'fornecedor', 'codigo_fornecedor', 'descricao_fornecedor')
                
                # Log dos resultados da separação
                if 'codigo_fornecedor' in df.columns and 'descricao_fornecedor' in df.columns:
                    logger.info(f"Separação concluída - Exemplos:")
                    sample_data = df[['fornecedor', 'codigo_fornecedor', 'descricao_fornecedor']].head(5)
                    for idx, row in sample_data.iterrows():
                        logger.info(f"  '{row['fornecedor']}' -> '{row['codigo_fornecedor']}' / '{row['descricao_fornecedor']}'")
            
            if 'titulo' in df.columns:
                logger.info("Limpando coluna 'titulo' - removendo letras, mantendo números e hífen")
                df['titulo'] = df['titulo'].astype(str).str.replace(r'[^0-9]', '', regex=True)
                
                # Log dos resultados da limpeza
                logger.info(f"Exemplos de títulos após limpeza: {df['titulo'].head(10).tolist()}")

            for date_col in ['data_emissao', 'data_vencimento']:
                if date_col in df.columns:
                    try:
                        # Converte para string primeiro para garantir consistência
                        df[date_col] = df[date_col].astype(str)
                        
                        # Remove espaços em branco
                        df[date_col] = df[date_col].str.strip()
                        
                        # Substitui valores vazios por NaN
                        df[date_col] = df[date_col].replace(['', 'nan', 'None', 'NaT'], np.nan)
                        
                        # Tenta converter para datetime
                        df[date_col] = pd.to_datetime(
                        df[date_col],
                        dayfirst=True,   # aceita tanto 01/09/2025 quanto 2025-09-01
                        errors='coerce'
                    )

                        
                        # Se não conseguir com formato específico, tenta inferir
                        if df[date_col].isna().any():
                            df[date_col] = df[date_col].dt.strftime('%Y-%m-%d')

                        
                        # Formata para string no formato brasileiro
                        df[date_col] = df[date_col].dt.strftime('%d/%m/%Y')
                        
                        # Substitui NaT por None
                        df[date_col] = df[date_col].replace('NaT', None)
                        
                    except Exception as e:
                        logger.warning(f"Erro ao converter {date_col}: {e}")
                        df[date_col] = None
            
            # Diagnóstico após a limpeza
            for date_col in ['data_emissao', 'data_vencimento']:
                if date_col in df.columns:
                    logger.info(f"Coluna {date_col} após limpeza:")
                    logger.info(f"  Tipo: {df[date_col].dtype}")
                    logger.info(f"  Primeiros valores: {df[date_col].head().tolist()}")
                    logger.info(f"  Não nulos: {df[date_col].notna().sum()}")
            
            # Remove registros de fornecedores NDF/PA
            # if 'fornecedor' in df.columns:
            #     mask = df['fornecedor'].str.contains(r'\bNDF\b|\bPA\b', case=False, na=False)
            #     logger.info(f"Removendo {mask.sum()} registros de NDF/PA/BOL/EMP/TX/INS/ISS/TXA/IRF")
            #     df = df[~mask]
            
            # Garante que todas as colunas obrigatórias existam
            required_cols = ['fornecedor', 'titulo', 'parcela', 'tipo_titulo', 
                            'data_emissao', 'data_vencimento', 'valor_original',
                            'saldo_devedor', 'situacao', 'conta_contabil', 'centro_custo']
            
            # Adiciona colunas de código e descrição se foram criadas
            if 'codigo_fornecedor' in df.columns:
                required_cols.append('codigo_fornecedor')
            if 'descricao_fornecedor' in df.columns:
                required_cols.append('descricao_fornecedor')
            
            for col in required_cols:
                if col not in df.columns:
                    df[col] = np.nan
            
            # Limpa e converte colunas numéricas
            num_cols = ['valor_original', 'saldo_devedor', 'titulos_vencer']  # ADICIONADO titulos_vencer
            for col in num_cols:
                if col in df.columns:
                    # Converte para string primeiro
                    df[col] = df[col].astype(str)
                    
                    # 1. Mantém apenas dígitos, vírgula, ponto e sinal
                    df[col] = df[col].str.replace(r'[^\d,.-]', '', regex=True)

                    # 2. Remove pontos de milhar
                    df[col] = df[col].str.replace('.', '', regex=False)

                    # 3. Troca vírgula por ponto (decimal BR → padrão Python)
                    df[col] = df[col].str.replace(',', '.', regex=False)

                    # 4. Converte para float
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

            logger.info(f"DataFrame final - shape: {df.shape}")
            logger.info(f"Colunas finais: {df.columns.tolist()}")
            
            return df
        except Exception as e:
            error_msg = f"Erro na limpeza de dados financeiros: {e}"
            logger.error(error_msg)
            raise InvalidDataFormat(error_msg, tipo_dado="financeiro") from e

    def _clean_modelo1_data(self, df):
        """
        Limpeza específica para dados do modelo1 (ctbr040).
        Aplica separação de código e descrição durante a importação.
        """
        try:
            # APLICA SEPARAÇÃO DE CÓDIGO E DESCRIÇÃO SE NECESSÁRIO
            if 'descricao_conta' in df.columns:
                # Tenta separar código e descrição da coluna descricao_conta
                df = self.separar_codigo_descricao(df, 'descricao_conta', 'codigo_fornecedor_temp', 'descricao_fornecedor_temp')
                
                # Se não temos código_fornecedor, usa o temporário
                if 'codigo_fornecedor' not in df.columns or df['codigo_fornecedor'].isna().all():
                    if 'codigo_fornecedor_temp' in df.columns:
                        df['codigo_fornecedor'] = df['codigo_fornecedor_temp']
                        df['descricao_fornecedor'] = df['descricao_fornecedor_temp']
                        df = df.drop(['codigo_fornecedor_temp', 'descricao_fornecedor_temp'], axis=1, errors='ignore')
                        logger.info("Código e descrição extraídos da coluna descricao_conta")

            # Classifica tipo de fornecedor baseado na descrição da conta
            if 'descricao_conta' in df.columns:
                df['tipo_fornecedor'] = df['descricao_conta'].apply(
                    lambda x: 'FORNECEDOR NACIONAL' if 'FORNEC' in str(x).upper() and 'NAC' in str(x).upper()
                    else 'FORNECEDOR' if 'FORNEC' in str(x).upper()
                    else 'OUTROS'
                )
            
            # Preenche códigos e descrições de fornecedor
            if 'codigo_fornecedor' not in df.columns:
                df['codigo_fornecedor'] = None
            if 'descricao_fornecedor' not in df.columns:
                df['descricao_fornecedor'] = None
            
            if 'codigo_fornecedor' in df.columns:
                df['codigo_fornecedor'] = df['codigo_fornecedor'].astype(str).str.strip()
                df['codigo_fornecedor'] = df['codigo_fornecedor'].str.replace(r'^(AF|F)', '', regex=True)

            # Extrai código do fornecedor da descrição da conta
            if df['codigo_fornecedor'].isna().all() and 'descricao_conta' in df.columns:
                df['codigo_fornecedor'] = df['descricao_conta'].str.extract(r'(\d{4,})', expand=False)
                df['descricao_fornecedor'] = df['descricao_conta']

            if 'descricao_conta' in df.columns:
                # Remove registros que contenham "OUTROS" na descrição
                antes = len(df)
                df = df[~df['descricao_conta'].str.contains('OUTROS', na=False)]
                depois = len(df)
                removidos = antes - depois
                if removidos > 0:
                    logger.info(f"Removidos {removidos} registros de 'FORNECEDORES OUTROS'")
            
            # Limpa e converte colunas numéricas
            num_cols = ['saldo_anterior', 'debito', 'credito', 'saldo_atual']
            for col in num_cols:
                if col in df.columns:
                    # Primeiro tenta aplicar a formatação de crédito se necessário
                    if col == 'credito':
                        df[col] = df[col].apply(self.formatar_credito)
                    else:
                        # Para outras colunas, usa conversão direta
                        df[col] = pd.to_numeric(
                            df[col].astype(str)
                            .str.replace(r'[^\d,-]', '', regex=True)
                            .str.replace(',', '.'),
                            errors='coerce'
                        ).fillna(0)
            
            return df
        except Exception as e:
            error_msg = f"Erro na limpeza de dados do modelo1: {e}"
            logger.error(error_msg)
            raise InvalidDataFormat(error_msg, tipo_dado="modelo1") from e

    def formatar_credito(self, valor):
        if pd.isna(valor):
            return None
        
        valor_str = str(valor).strip()

        # Verifica se é crédito (C) ou débito (D)
        is_credito = valor_str.endswith("C")
        is_debito = valor_str.endswith("D")

        # Remove tudo que não for dígito, vírgula ou ponto
        valor_str = re.sub(r'[^\d,]', '', valor_str)

        try:
            # CORREÇÃO: Remover pontos de milhar ANTES de converter
            # Se tiver vírgula, assume que é decimal
            if ',' in valor_str:
                # Remove pontos de milhar e troca vírgula por ponto
                valor_str = valor_str.replace('.', '').replace(',', '.')
            else:
                # Se não tem vírgula, pode ser que já esteja no formato correto
                # Mas remove pontos de milhar para segurança
                valor_str = valor_str.replace('.', '')
            
            valor_float = float(valor_str)
            
            # Ajusta o sinal baseado no tipo (C ou D)
            if is_credito:
                valor_float = -abs(valor_float)  # Crédito é negativo
            elif is_debito:
                valor_float = abs(valor_float)   # Débito é positivo
            
        except Exception as e:
            logger.warning(f"Erro ao converter valor '{valor}': {e}")
            valor_float = 0.0

        return valor_float  # Retorna o valor numérico, não formatado


    def _clean_contas_itens_data(self, df):
        """
        Limpeza específica para dados de contas_itens.
        Aplica separação de código e descrição durante a importação.
        """
        try:
            # APLICA SEPARAÇÃO DE CÓDIGO E DESCRIÇÃO SE NECESSÁRIO
            if 'descricao_item' in df.columns:
                df = self.separar_codigo_descricao(df, 'descricao_item', 'codigo_fornecedor_temp', 'descricao_fornecedor_temp')
                
                # Se não temos código_fornecedor, usa o temporário
                if 'codigo_fornecedor' not in df.columns or df['codigo_fornecedor'].isna().all():
                    if 'codigo_fornecedor_temp' in df.columns:
                        df['codigo_fornecedor'] = df['codigo_fornecedor_temp']
                        df['descricao_fornecedor'] = df['descricao_fornecedor_temp']
                        df = df.drop(['codigo_fornecedor_temp', 'descricao_fornecedor_temp'], axis=1, errors='ignore')
                        logger.info("Código e descrição extraídos da coluna descricao_item")

            if 'codigo_fornecedor' in df.columns:
                df['codigo_fornecedor'] = df['codigo_fornecedor'].astype(str).str.strip()
                df['codigo_fornecedor'] = df['codigo_fornecedor'].str.replace(r'^(AF|F)', '', regex=True)

            reverse_mapping = {v: k for k, v in self.settings.COLUNAS_CONTAS_ITENS.items()}
            df = df.rename(columns=reverse_mapping)
            
            for col in self.settings.COLUNAS_CONTAS_ITENS.keys():
                if col not in df.columns:
                    df[col] = None

            if 'credito' in df.columns:
                df['credito'] = df['credito'].apply(self.formatar_credito)

            if 'saldo_anterior' in df.columns:
                df['saldo_anterior'] = df['saldo_anterior'].apply(self.formatar_credito)

            if 'saldo_atual' in df.columns:
                df['saldo_atual'] = df['saldo_atual'].apply(self.formatar_credito)

            return df
        except Exception as e:
            logger.error(f"Erro ao limpar dados de Contas x Itens: {e}")
            raise


    def _clean_adiantamento_data(self, df):
        """
        Limpeza específica para dados de adiantamentos.
        Aplica separação de código e descrição durante a importação.
        """
        try:
            # APLICA SEPARAÇÃO DE CÓDIGO E DESCRIÇÃO SE NECESSÁRIO
            if 'descricao_item' in df.columns:
                df = self.separar_codigo_descricao(df, 'descricao_item', 'codigo_fornecedor_temp', 'descricao_fornecedor_temp')
                
                # Se não temos código_fornecedor, usa o temporário
                if 'codigo_fornecedor' not in df.columns or df['codigo_fornecedor'].isna().all():
                    if 'codigo_fornecedor_temp' in df.columns:
                        df['codigo_fornecedor'] = df['codigo_fornecedor_temp']
                        df['descricao_fornecedor'] = df['descricao_fornecedor_temp']
                        df = df.drop(['codigo_fornecedor_temp', 'descricao_fornecedor_temp'], axis=1, errors='ignore')
                        logger.info("Código e descrição extraídos da coluna descricao_item (adiantamento)")

            if 'codigo_fornecedor' in df.columns:
                df['codigo_fornecedor'] = df['codigo_fornecedor'].astype(str).str.strip()
                df['codigo_fornecedor'] = df['codigo_fornecedor'].str.replace(r'^(AF|F)', '', regex=True)

            reverse_mapping = {v: k for k, v in self.settings.COLUNAS_ADIANTAMENTO.items()}
            df = df.rename(columns=reverse_mapping)
            
            # Garante que todas as colunas existam
            for col in self.settings.COLUNAS_ADIANTAMENTO.keys():
                if col not in df.columns:
                    df[col] = None

            # Aplica a formatação de crédito nas mesmas colunas
            if 'credito' in df.columns:
                df['credito'] = df['credito'].apply(self.formatar_credito)

            if 'saldo_anterior' in df.columns:
                df['saldo_anterior'] = df['saldo_anterior'].apply(self.formatar_credito)

            if 'saldo_atual' in df.columns:
                df['saldo_atual'] = df['saldo_atual'].apply(self.formatar_credito)

            return df[list(self.settings.COLUNAS_ADIANTAMENTO.keys())]
        except Exception as e:
            logger.error(f"Erro ao limpar dados de Adiantamentos: {e}")
            raise

    def _get_column_mapping(self, file_path: Path):
        """
        Retorna mapeamento de colunas baseado no nome do arquivo e extensão.
        """
        filename = file_path.stem.lower()
        ext = file_path.suffix.lower()
        
        # Mapeamento para arquivos XLSX
        if ext == ".xlsx":
            if 'finr150' in filename:
                return getattr(self.settings, 'COLUNAS_FINANCEIRO', {})
            elif 'ctbr040' in filename:
                return getattr(self.settings, 'COLUNAS_MODELO1', {})
            elif 'ctbr140' in filename:
                return getattr(self.settings, 'COLUNAS_CONTAS_ITENS', {})
            elif 'ctbr100' in filename:  
                return getattr(self.settings, 'COLUNAS_ADIANTAMENTO', {})
        
        # Mapeamento para arquivos TXT
        elif ext == ".txt":
            if 'ctbr140' in filename:
                return getattr(self.settings, 'COLUNAS_CONTAS_ITENS', {})
            elif 'ctbr100' in filename:
                return getattr(self.settings, 'COLUNAS_ADIANTAMENTO', {})
        
        # Mapeamento para arquivos XML
        elif ext == ".xml":
            if 'ctbr140' in filename:
                return getattr(self.settings, 'COLUNAS_CONTAS_ITENS', {})
            elif 'ctbr100' in filename:
                return getattr(self.settings, 'COLUNAS_ADIANTAMENTO', {})
        
        # Se não encontrou mapeamento, retorna dicionário vazio
        logger.warning(f"Tipo de planilha não reconhecido: {file_path.name}")
        return {}

    def process_data(self):
        """
        Processa os dados importados e gera resultados da conciliação.
        """
        try:
            self.conn.execute("BEGIN TRANSACTION")
            cursor = self.conn.cursor()
            
            # Obtém período de referência
            data_inicial, data_final = self._get_datas_referencia()
            
            # Limpa tabelas de resultados anteriores
            cursor.execute(f"DELETE FROM {self.settings.TABLE_RESULTADO}")
            cursor.execute(f"DELETE FROM {self.settings.TABLE_RESULTADO_ADIANTAMENTO}")

            #  CORREÇÃO: Query consolidada para fornecedores do financeiro (NF/FT)
            query_consolidada = f"""
                INSERT INTO {self.settings.TABLE_RESULTADO}
                (codigo_fornecedor, descricao_fornecedor, saldo_financeiro, saldo_contabil, status)
                
                -- Busca fornecedores do financeiro (NF/FT)
                SELECT 
                    COALESCE(NULLIF(TRIM(f.codigo_fornecedor), ''), TRIM(f.fornecedor)) as codigo_fornecedor,
                    COALESCE(NULLIF(TRIM(f.descricao_fornecedor), ''), TRIM(f.fornecedor)) as descricao_fornecedor,
                    SUM(COALESCE(f.valor_original, 0)) as saldo_financeiro,
                    0 as saldo_contabil,  -- Será atualizado depois
                    'Pendente' as status
                FROM 
                    {self.settings.TABLE_FINANCEIRO} f
                WHERE 
                    f.excluido = 0
                    AND UPPER(f.tipo_titulo) IN ('NF','FT')
                GROUP BY 
                    COALESCE(NULLIF(TRIM(f.codigo_fornecedor), ''), TRIM(f.fornecedor)),
                    COALESCE(NULLIF(TRIM(f.descricao_fornecedor), ''), TRIM(f.fornecedor))
            """
            cursor.execute(query_consolidada)

            #  CORREÇÃO CRÍTICA: Atualiza com valores contábeis APENAS da conta correta (Fornecedores Nacionais)
            query_atualiza_contabil = f"""
                UPDATE {self.settings.TABLE_RESULTADO}
                SET 
                    saldo_contabil = (
                        SELECT COALESCE(SUM(ci.saldo_atual), 0)
                        FROM {self.settings.TABLE_CONTAS_ITENS} ci
                        WHERE 
                            ci.conta_contabil LIKE '2.01.02.01.0001%'
                            AND REPLACE(REPLACE(UPPER(TRIM(ci.codigo_fornecedor)), 'AF', ''), 'F', '') =
                                REPLACE(REPLACE(UPPER(TRIM({self.settings.TABLE_RESULTADO}.codigo_fornecedor)), 'AF', ''), 'F', '')
                            AND ci.codigo_fornecedor IS NOT NULL
                            AND ci.codigo_fornecedor != ''
                    ),
                    detalhes = (
                        SELECT GROUP_CONCAT(
                            'Conta: ' || ci.conta_contabil || 
                            ' | Item: ' || ci.descricao_item || 
                            ' | Valor: R$ ' || ROUND(COALESCE(ci.saldo_atual, 0), 2), ' | '
                        )
                        FROM {self.settings.TABLE_CONTAS_ITENS} ci
                        WHERE 
                            ci.conta_contabil LIKE '2.01.02.01.0001%'
                            AND REPLACE(REPLACE(UPPER(TRIM(ci.codigo_fornecedor)), 'AF', ''), 'F', '') =
                                REPLACE(REPLACE(UPPER(TRIM({self.settings.TABLE_RESULTADO}.codigo_fornecedor)), 'AF', ''), 'F', '')
                            AND ci.codigo_fornecedor IS NOT NULL
                            AND ci.codigo_fornecedor != ''
                    )
                WHERE EXISTS (
                    SELECT 1
                    FROM {self.settings.TABLE_CONTAS_ITENS} ci2
                    WHERE 
                        ci2.conta_contabil LIKE '2.01.02.01.0001%'
                        AND REPLACE(REPLACE(UPPER(TRIM(ci2.codigo_fornecedor)), 'AF', ''), 'F', '') =
                            REPLACE(REPLACE(UPPER(TRIM({self.settings.TABLE_RESULTADO}.codigo_fornecedor)), 'AF', ''), 'F', '')
                        AND ci2.codigo_fornecedor IS NOT NULL
                        AND ci2.codigo_fornecedor != ''
                )
            """


            cursor.execute(query_atualiza_contabil)

            #  CORREÇÃO: Adiciona adiantamentos aos saldos contábeis (conta específica)
            query_adiantamento = f"""
                UPDATE {self.settings.TABLE_RESULTADO}
                SET 
                    saldo_contabil = saldo_contabil + (
                        SELECT COALESCE(SUM(saldo_atual), 0)
                        FROM {self.settings.TABLE_ADIANTAMENTO} a
                        WHERE 
                            --  FILTRO DA CONTA DE ADIANTAMENTO
                            a.conta_contabil LIKE '1.01.06.02.0001%'
                            AND a.codigo_fornecedor = {self.settings.TABLE_RESULTADO}.codigo_fornecedor
                    )
                WHERE EXISTS (
                    SELECT 1
                    FROM {self.settings.TABLE_ADIANTAMENTO} a2
                    WHERE 
                        a2.conta_contabil LIKE '1.01.06.02.0001%'
                        AND a2.codigo_fornecedor = {self.settings.TABLE_RESULTADO}.codigo_fornecedor
                )
            """
            cursor.execute(query_adiantamento)

            #  CORREÇÃO: Insere fornecedores contábeis que não existem no financeiro
            query_contabeis_faltantes = f"""
                INSERT INTO {self.settings.TABLE_RESULTADO}
                (codigo_fornecedor, descricao_fornecedor, saldo_financeiro, saldo_contabil, status)
                
                SELECT 
                    COALESCE(NULLIF(TRIM(ci.codigo_fornecedor), ''), ci.descricao_fornecedor) as codigo_fornecedor,
                    COALESCE(NULLIF(TRIM(ci.descricao_fornecedor), ''), ci.descricao_item) as descricao_fornecedor,
                    0 as saldo_financeiro,
                    SUM(COALESCE(ci.saldo_atual, 0)) as saldo_contabil,
                    'Pendente' as status
                FROM 
                    {self.settings.TABLE_CONTAS_ITENS} ci
                WHERE 
                    ci.conta_contabil LIKE '2.01.02.01.0001%'
                    AND NOT EXISTS (
                        SELECT 1
                        FROM {self.settings.TABLE_RESULTADO} r
                        WHERE r.codigo_fornecedor = COALESCE(NULLIF(TRIM(ci.codigo_fornecedor), ''), ci.descricao_fornecedor)
                    )
                GROUP BY 
                    COALESCE(NULLIF(TRIM(ci.codigo_fornecedor), ''), ci.descricao_fornecedor),
                    COALESCE(NULLIF(TRIM(ci.descricao_fornecedor), ''), ci.descricao_item)
            """
            cursor.execute(query_contabeis_faltantes)
            
            # Cálculo de diferenças e status
            query_diferenca = f"""
                UPDATE {self.settings.TABLE_RESULTADO}
                SET 
                    diferenca = ROUND(COALESCE(saldo_contabil, 0) - COALESCE(saldo_financeiro, 0), 2),
                    status = CASE 
                        WHEN saldo_contabil IS NULL AND saldo_financeiro IS NULL THEN 'Pendente'
                        WHEN ABS(COALESCE(saldo_financeiro, 0) - COALESCE(saldo_contabil, 0)) <= 
                            (0.03 * CASE 
                                WHEN ABS(COALESCE(saldo_contabil, 0)) > ABS(COALESCE(saldo_financeiro, 0)) 
                                THEN ABS(COALESCE(saldo_contabil, 0)) 
                                ELSE ABS(COALESCE(saldo_financeiro, 0)) 
                            END)
                            THEN 'Conferido' 
                        ELSE 'Divergente' 
                    END
            """
            cursor.execute(query_diferenca)
            
            # Query para investigação de divergências
            query_investigacao = f"""
                UPDATE {self.settings.TABLE_RESULTADO}
                SET detalhes = COALESCE(detalhes, '') || 
                    ' | Divergência: R$ ' || ABS(diferenca) || 
                    '. Itens Contábeis encontrados: ' || 
                    COALESCE(
                        (SELECT COUNT(*) || ' itens'
                        FROM {self.settings.TABLE_CONTAS_ITENS} ci
                        WHERE (ci.codigo_fornecedor = {self.settings.TABLE_RESULTADO}.codigo_fornecedor 
                                OR ci.descricao_fornecedor = {self.settings.TABLE_RESULTADO}.descricao_fornecedor)
                        AND ci.conta_contabil LIKE '2.01.02.01.0001%'),
                        'Nenhum item específico encontrado'
                    )
                WHERE status = 'Divergente'
            """
            cursor.execute(query_investigacao)
            
            # Para fornecedores divergentes sem itens específicos
            cursor.execute(f"""
                UPDATE {self.settings.TABLE_RESULTADO}
                SET detalhes = 'Divergência: R$ ' || ABS(diferenca) || 
                            '. Investigar manualmente no sistema. Nenhum item contábil específico encontrado para análise automática.'
                WHERE status = 'Divergente' 
                AND (detalhes IS NULL OR detalhes = '')
            """)
            
            # Classifica por ordem de importância
            try:
                cursor.execute(f"""
                    UPDATE {self.settings.TABLE_RESULTADO}
                    SET ordem_importancia = (
                        SELECT COUNT(*) 
                        FROM {self.settings.TABLE_RESULTADO} r2 
                        WHERE ABS(COALESCE(r2.diferenca, 0)) >= ABS(COALESCE({self.settings.TABLE_RESULTADO}.diferenca, 0))
                    )
                """)
            except Exception as rank_error:
                logger.error(f"Erro ao classificar por importância: {rank_error}")

            # Atualiza detalhes para registros não divergentes
            cursor.execute(f"""
                UPDATE {self.settings.TABLE_RESULTADO}
                SET detalhes = 
                    CASE 
                        WHEN status = 'Conferido' THEN 'Conciliação dentro da tolerância'
                        WHEN status = 'Pendente' THEN 'Financeiro: R$ ' || COALESCE(saldo_financeiro, 0) || 
                                                    ' | Contábil: R$ ' || COALESCE(saldo_contabil, 0) || 
                                                    ' | Diferença: R$ ' || COALESCE(diferenca, 0)
                        ELSE detalhes  -- Mantém os detalhes da investigação para divergências
                    END
                WHERE detalhes IS NULL OR detalhes = ''
            """)
            
            # Processamento de adiantamentos
            self._process_adiantamentos()

            self.conn.commit()
            logger.info("Processamento de dados concluído com sucesso - filtros de conta contábil aplicados")
            return True
            
        except Exception as e:
            error_msg = f"Erro ao processar dados: {e}"
            logger.error(error_msg, exc_info=True)
            self.conn.rollback()
            raise ExcecaoNaoMapeadaError(error_msg) from e
    def _get_datas_referencia(self, data_referencia=None):
        """
        Calcula as datas inicial e final para o relatório Contas X Itens
        conforme as regras especificadas.
        """
        if data_referencia is None:
            data_referencia = datetime.now()
        
        dia = data_referencia.day
        mes = data_referencia.month
        ano = data_referencia.year
        
        # Verifica se é o último dia do mês
        ultimo_dia_mes = calendar.monthrange(ano, mes)[1]
        eh_ultimo_dia = dia == ultimo_dia_mes
        
        # INICIALIZA AS VARIÁVEIS COM VALORES PADRÃO
        data_inicial = data_referencia  # valor padrão
        data_final = data_referencia    # valor padrão
        
        if eh_ultimo_dia:
            # Regra para último dia do mês
            # Data Inicial: primeiro dia do mês anterior
            if mes == 1:
                data_inicial = datetime(ano - 1, 12, 1)
            else:
                data_inicial = datetime(ano, mes - 1, 1)
            
            # Data Final: último dia do mês anterior
            if mes == 1:
                ultimo_dia_anterior = calendar.monthrange(ano - 1, 12)[1]
                data_final = datetime(ano - 1, 12, ultimo_dia_anterior)
            else:
                ultimo_dia_anterior = calendar.monthrange(ano, mes - 1)[1]
                data_final = datetime(ano, mes - 1, ultimo_dia_anterior)
        else:
            # SE NÃO FOR ÚLTIMO DIA DO MÊS, USA DATAS PADRÃO
            # Por exemplo: primeiro dia do mês atual até hoje
            data_inicial = datetime(ano, mes, 1)
            data_final = data_referencia
        
        # Formata as datas
        data_inicial_str = data_inicial.strftime('%d/%m/%Y')
        data_final_str = data_final.strftime('%d/%m/%Y')
        
        return data_inicial_str, data_final_str
    
    def validate_data_consistency(self):
        try:
            cursor = self.conn.cursor()
            
            # Verifica totais financeiros vs contábeis
            query = f"""
                SELECT 
                    (SELECT SUM(saldo_devedor) FROM {self.settings.TABLE_FINANCEIRO} 
                    WHERE excluido = 0 AND UPPER(tipo_titulo) NOT IN ('NDF', 'PA', 'BOL', 'EMP', 'TX', 'INS', 'ISS', 'TXA', 'IRF')) as total_financeiro,
                    (SELECT SUM(saldo_atual) FROM {self.settings.TABLE_MODELO1} 
                    WHERE descricao_conta LIKE 'FORNECEDOR%') as total_contabil
            """
            cursor.execute(query)
            totals = cursor.fetchone()
            
            diferenca_percentual = abs(totals[0] - totals[1]) / max(totals[0], totals[1]) * 100
            
            if diferenca_percentual > 5:  # Tolerância de 5% para o total
                logger.warning(f"Diferença significativa entre totais: Financeiro={totals[0]}, Contábil={totals[1]}")
            
            return True
        except Exception as e:
            error_msg = f"Erro na validação de consistência: {e}"
            logger.error(error_msg)
            return False

    def _apply_metadata_styles(self, worksheet, metadata_items, metadata_values):
        """
        Aplica estilos à aba de metadados de forma otimizada.
        """
        try:
            # Define estilos
            title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_font = Font(color="FFFFFF", bold=True, size=14)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Formata título (primeira linha)
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = title_fill
                    cell.font = title_font
                    cell.border = thin_border
            
            # Formata cabeçalho (segunda linha)
            for row in worksheet.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
            
            # Aplica bordas a todas as células restantes
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
            
            # Ajusta largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)  # Limita a largura máxima
                worksheet.column_dimensions[column_letter].width = adjusted_width

            for row_idx, (item, value) in enumerate(zip(metadata_items, metadata_values), 1):
                if '---' in str(item) or '---' in str(value):
                    # Aplica fundo cinza para separadores
                    for col in range(1, 3):
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        
        except Exception as e:
            logger.warning(f"Erro ao aplicar estilos de metadados: {e}")
            
    def _apply_styles(self, worksheet):
        """
        Aplica estilos visuais básicos à planilha Excel de forma otimizada.
        """
        try:
            # Define estilos
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            align_center = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Aplica estilos ao cabeçalho em lote
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = thin_border
            
            # Identifica colunas monetárias
            header = [c.value for c in worksheet[1] if c.value is not None]
            monetary_headers = {
                'Valor Financeiro', 'Valor Contábil', 'Diferença',
                'Valor em Aberto', 'Valor Provisionado', 'Saldo Atual',
                'Débito', 'Crédito', 'Valor Original', 'Saldo Devedor',
                'Quantidade', 'Valor Unitário', 'Valor Total'
            }
            
            # Aplica formatação monetária em colunas inteiras 
            for col_idx, col_name in enumerate(header, 1):
                if col_name in monetary_headers:
                    col_letter = get_column_letter(col_idx)
                    for cell in worksheet[col_letter][1:]:  # Pula o cabeçalho
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = 'R$ #,##0.00;[Red]R$ -#,##0.00'
            
            # Aplica bordas a todas as células (em lote por linha)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
            
            # Ajusta largura das colunas automaticamente
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)  # Limita a largura máxima
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"Erro ao aplicar estilos básicos: {e}")

    def _apply_enhanced_styles(self, worksheet, stats):
        """
        Aplica estilos visuais melhorados com formatação otimizada.
        """
        try:
            # Cores para formatação condicional (DEFINIR NO INÍCIO DO MÉTODO)
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            red_font = Font(color="9C0006", bold=True)
            
            # Define estilos
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            align_center = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Aplica estilos ao cabeçalho em lote
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = thin_border
            
            # Identifica índices de colunas para formatação monetária
            header = [c.value for c in worksheet[1] if c.value is not None]
            monetary_columns = [
                'Valor Financeiro', 'Valor Contábil', 'Diferença',
                'Valor em Aberto', 'Valor Provisionado', 'Saldo Atual',
                'Débito', 'Crédito'
            ]
            
            # Aplica formatação monetária em colunas inteiras
            for col_idx, col_name in enumerate(header, 1):
                if col_name in monetary_columns:
                    col_letter = get_column_letter(col_idx)
                    for cell in worksheet[col_letter][1:]:  # Pula o cabeçalho
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = 'R$ #,##0.00;[Red]R$ -#,##0.00'
            
            # Identifica colunas de status e diferença
            status_idx = header.index("Status") + 1 if "Status" in header else None
            diferenca_idx = header.index("Diferença") + 1 if "Diferença" in header else None
            if diferenca_idx:
                diff_cell = row[diferenca_idx-1]
                if diff_cell.value is not None and diff_cell.value != 0:
                    diff_cell.fill = red_fill  # ou amarelo_fill, se preferir apenas destacar

            # Cores para formatação condicional
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            red_font = Font(color="9C0006", bold=True)
            
            # Aplica formatação condicional por linhas (mais eficiente)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                # Formata valores negativos em vermelho (apenas para coluna Diferença)
                if diferenca_idx:
                    diff_cell = row[diferenca_idx-1]
                    if diff_cell.value is not None and diff_cell.value != 0:
                        diff_cell.fill = yellow_fill
                        diff_cell.font = red_font

                # Formatação baseada no status
                if status_idx:
                    status_cell = row[status_idx-1]  # -1 porque index começa em 0
                    status_value = status_cell.value if status_cell.value else ""
                    
                    fill_color = None
                    if status_value == 'Conferido':
                        fill_color = green_fill
                    elif status_value == 'Divergente':
                        fill_color = red_fill
                    elif status_value == 'Pendente':
                        fill_color = yellow_fill
                    
                    # Aplica o preenchimento apenas se necessário
                    if fill_color:
                        for cell in row:
                            cell.fill = fill_color
                
                # Aplica bordas a todas as células
                for cell in row:
                    cell.border = thin_border
            
            # Ajusta largura das colunas automaticamente
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2) * 1.2, 50)  # Limita a largura máxima
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"Erro ao aplicar estilos avançados: {e}")

    def _optimize_worksheet_performance(self, worksheet):
        """
        Otimiza a planilha para melhor performance de escrita.
        """
        # Desativa propriedades que tornam a escrita lenta
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_view.showRowColHeaders = False

    def _protect_sheets(self, workbook):
        """
        Protege todas as abas, exceto a coluna 'Observações' na aba de resumo.
        """
        try:
            from openpyxl.worksheet.protection import SheetProtection
            
            for sheetname in workbook.sheetnames:
                sheet = workbook[sheetname]
                if sheet is not None:
                    # Protege a planilha inteira
                    sheet.protection = SheetProtection(
                        sheet=True, 
                        selectLockedCells=False,
                        selectUnlockedCells=False
                    )
                    
                    # Libera apenas a coluna "Observações" na aba de resumo
                    if sheetname == 'Resumo da Conciliação':
                        header = [cell.value for cell in sheet[1] if cell.value is not None]
                        if "Observações" in header:
                            obs_col_idx = header.index("Observações") + 1
                            for row in range(2, sheet.max_row + 1):
                                cell = sheet.cell(row=row, column=obs_col_idx)
                                cell.protection = Protection(locked=False)
        
        except Exception as e:
            logger.warning(f"")

    def _process_adiantamentos(self):
        """Processa especificamente a conciliação de adiantamentos para a planilha separada"""
        try:
            cursor = self.conn.cursor()
            
            # Limpa a tabela de resultado_adiantamento
            cursor.execute(f"DELETE FROM {self.settings.TABLE_RESULTADO_ADIANTAMENTO}")
            
            #  CORREÇÃO: Verifica se as colunas já existem antes de tentar adicioná-las
            cursor.execute(f"PRAGMA table_info({self.settings.TABLE_FINANCEIRO})")
            existing_columns = [col[1] for col in cursor.fetchall()]
            
            if 'codigo_fornecedor' not in existing_columns:
                cursor.execute("ALTER TABLE financeiro ADD COLUMN codigo_fornecedor TEXT")
            
            if 'descricao_fornecedor' not in existing_columns:
                cursor.execute("ALTER TABLE financeiro ADD COLUMN descricao_fornecedor TEXT")
            
            # Atualiza as colunas com valores padrão se estiverem vazias
            cursor.execute("UPDATE financeiro SET codigo_fornecedor = fornecedor WHERE codigo_fornecedor IS NULL")
            cursor.execute("UPDATE financeiro SET descricao_fornecedor = fornecedor WHERE descricao_fornecedor IS NULL")

            #  CORREÇÃO: Insere dados financeiros de adiantamentos (NDF, PA)
            query_adiantamento_financeiro = f"""
                INSERT INTO {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
                (codigo_fornecedor, descricao_fornecedor, total_financeiro, status)
                SELECT 
                    COALESCE(NULLIF(TRIM(f.codigo_fornecedor), ''), TRIM(f.fornecedor)) as codigo_fornecedor,
                    COALESCE(NULLIF(TRIM(f.descricao_fornecedor), ''), TRIM(f.fornecedor)) as descricao_fornecedor,
                    SUM(COALESCE(tit_vencidos_valor_nominal, 0) + COALESCE(titulos_a_vencer_valor_nominal, 0)) as total_financeiro, 
                    'Pendente' as status
                FROM 
                    {self.settings.TABLE_FINANCEIRO} f
                WHERE 
                    excluido = 0
                    AND UPPER(tipo_titulo) IN ('NDF', 'PA')
                GROUP BY 
                    COALESCE(NULLIF(TRIM(f.codigo_fornecedor), ''), TRIM(f.fornecedor)),
                    COALESCE(NULLIF(TRIM(f.descricao_fornecedor), ''), TRIM(f.fornecedor))
            """
            cursor.execute(query_adiantamento_financeiro)

            #  CORREÇÃO: Atualiza com dados contábeis de adiantamento
            query_contabil_update = f"""
                UPDATE {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
                SET 
                    total_contabil = (
                        SELECT COALESCE(SUM(saldo_atual), 0)
                        FROM {self.settings.TABLE_ADIANTAMENTO} a
                        WHERE a.codigo_fornecedor = {self.settings.TABLE_RESULTADO_ADIANTAMENTO}.codigo_fornecedor
                        OR TRIM(a.descricao_fornecedor) = {self.settings.TABLE_RESULTADO_ADIANTAMENTO}.descricao_fornecedor
                    ),
                    detalhes = 'Adiantamento: ' || COALESCE((
                        SELECT GROUP_CONCAT(a2.descricao_fornecedor || ': R$ ' || a2.saldo_atual, ' | ')
                        FROM {self.settings.TABLE_ADIANTAMENTO} a2
                        WHERE a2.codigo_fornecedor = {self.settings.TABLE_RESULTADO_ADIANTAMENTO}.codigo_fornecedor
                        OR TRIM(a2.descricao_fornecedor) = {self.settings.TABLE_RESULTADO_ADIANTAMENTO}.descricao_fornecedor
                    ), 'Nenhum registro contábil')
                WHERE EXISTS (
                    SELECT 1
                    FROM {self.settings.TABLE_ADIANTAMENTO} a3
                    WHERE a3.codigo_fornecedor = {self.settings.TABLE_RESULTADO_ADIANTAMENTO}.codigo_fornecedor
                    OR TRIM(a3.descricao_fornecedor) = {self.settings.TABLE_RESULTADO_ADIANTAMENTO}.descricao_fornecedor
                )
            """
            cursor.execute(query_contabil_update)
            
            #  CORREÇÃO: Insere adiantamentos contábeis que não tiveram match financeiro
            query_contabeis_sem_match = f"""
                INSERT INTO {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
                (codigo_fornecedor, descricao_fornecedor, total_contabil, status, detalhes)
                SELECT 
                    a.codigo_fornecedor,
                    a.descricao_fornecedor,
                    SUM(COALESCE(a.saldo_atual, 0)) as total_contabil,
                    'Pendente' as status,
                    'Adiantamento contábil sem correspondência financeira' as detalhes
                FROM 
                    {self.settings.TABLE_ADIANTAMENTO} a
                WHERE 
                    a.codigo_fornecedor IS NOT NULL 
                    AND a.codigo_fornecedor <> ''
                    AND NOT EXISTS (
                        SELECT 1
                        FROM {self.settings.TABLE_RESULTADO_ADIANTAMENTO} r
                        WHERE r.codigo_fornecedor = a.codigo_fornecedor
                        OR TRIM(r.descricao_fornecedor) = TRIM(a.descricao_fornecedor)
                    )
                GROUP BY 
                    a.codigo_fornecedor, a.descricao_fornecedor
            """
            cursor.execute(query_contabeis_sem_match)
            
            
            query_diferenca = f"""
                UPDATE {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
                SET 
                    diferenca = ROUND(COALESCE(total_financeiro, 0) + COALESCE(total_contabil, 0), 2),
                    status = CASE 
                        WHEN total_contabil IS NULL AND total_financeiro IS NULL THEN 'Pendente'
                        WHEN total_contabil IS NULL AND total_financeiro IS NOT NULL THEN 'Divergente'
                        WHEN total_financeiro IS NULL AND total_contabil IS NOT NULL THEN 'Divergente'
                        WHEN ABS(COALESCE(total_financeiro, 0) + COALESCE(total_contabil, 0)) <= 
                            (0.03 * CASE 
                                WHEN ABS(COALESCE(total_contabil, 0)) > ABS(COALESCE(total_financeiro, 0)) 
                                THEN ABS(COALESCE(total_contabil, 0)) 
                                ELSE ABS(COALESCE(total_financeiro, 0)) 
                            END)
                            THEN 'Conferido' 
                        ELSE 'Divergente' 
                    END
            """
            cursor.execute(query_diferenca)
            
            #  CORREÇÃO: Atualiza detalhes para registros divergentes
            cursor.execute(f"""
                UPDATE {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
                SET detalhes = 
                    CASE 
                        WHEN status = 'Conferido' THEN 'Adiantamento conciliado'
                        WHEN status = 'Divergente' AND total_financeiro IS NULL THEN 
                            'Adiantamento contábil sem lançamento financeiro: R$ ' || COALESCE(total_contabil, 0)
                        WHEN status = 'Divergente' AND total_contabil IS NULL THEN 
                            'Adiantamento financeiro sem lançamento contábil: R$ ' || COALESCE(total_financeiro, 0)
                        ELSE 'Diferença: R$ ' || ABS(COALESCE(diferenca, 0)) || 
                            ' | Financeiro: R$ ' || COALESCE(total_financeiro, 0) ||
                            ' | Contábil: R$ ' || COALESCE(total_contabil, 0)
                    END
                WHERE detalhes IS NULL OR detalhes = ''
            """)
            
            logger.info("Processamento de adiantamentos concluído com sucesso")
            
        except Exception as e:
            error_msg = f"Erro no processamento de adiantamentos: {e}"
            logger.error(error_msg)
            raise

    def _recreate_adiantamento_table(self):
        """Recria a tabela de resultado_adiantamento com estrutura correta"""
        try:
            cursor = self.conn.cursor()
            cursor.execute(f"DROP TABLE IF EXISTS {self.settings.TABLE_RESULTADO_ADIANTAMENTO}")
            
            cursor.execute(f"""
                CREATE TABLE {self.settings.TABLE_RESULTADO_ADIANTAMENTO} (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    codigo_fornecedor TEXT,
                    descricao_fornecedor TEXT,
                    total_financeiro REAL DEFAULT 0,
                    total_contabil REAL DEFAULT 0,
                    diferenca REAL DEFAULT 0,
                    status TEXT CHECK(status IN ('Conferido', 'Divergente', 'Pendente')),
                    detalhes TEXT,
                    data_processamento TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            self.conn.commit()
            logger.info("Tabela resultado_adiantamento recriada com estrutura correta")
        except Exception as e:
            error_msg = f"Erro ao recriar tabela resultado_adiantamento: {e}"
            logger.error(error_msg)
            raise

    def separar_codigo_descricao(self, df, coluna_origem="Fornecedor", col_codigo="Codigo", col_descricao="Descricao"):
        """
        Versão MELHORADA: Extrai todos os dígitos do início da string para a coluna de código
        e remove esses dígitos da descrição.
        """
        if coluna_origem in df.columns:
            df = df.copy()
            df[col_codigo] = ""
            df[col_descricao] = ""
            
            for idx, valor in df[coluna_origem].items():
                if pd.notna(valor) and str(valor).strip() != "":
                    valor_str = str(valor).strip()
                    
                    # REGEX MELHORADA: Pega TODOS os dígitos do início, incluindo possíveis separadores
                    # Captura: 123, 123-456, 123.456, 123 456, etc.
                    match = re.match(r'^(\d+[\s\-\.\/]*\d*)', valor_str)
                    
                    if match:
                        codigo = match.group(1).strip()
                        # Remove qualquer caractere não numérico do código (opcional)
                        codigo_limpo = re.sub(r'[^\d]', '', codigo)  # Mantém apenas dígitos
                        # Ou mantenha com os separadores: codigo_limpo = codigo
                        
                        df.at[idx, col_codigo] = codigo_limpo
                        # Remove o código encontrado do início da descrição
                        descricao = valor_str[len(codigo):].strip()
                        # Remove hífens, pontos ou espaços extras no início
                        descricao = re.sub(r'^[\s\-\.\/]+', '', descricao)
                        df.at[idx, col_descricao] = descricao
                    else:
                        # Se não tem números no início, coloca vazio no código
                        df.at[idx, col_codigo] = ""
                        df.at[idx, col_descricao] = valor_str
                    
                    # Log para debug (opcional)
                    logger.debug(f"'{valor_str}' -> '{df.at[idx, col_codigo]}' / '{df.at[idx, col_descricao]}'")
            
            return df
    
    def export_to_excel(self, export_type="all"):
        """
        Exporta resultados para arquivo(s) Excel formatado(s) com metadados.
        
        Args:
            export_type: Tipo de exportação - "all", "fornecedores", "adiantamentos"
        """
        data_inicial_str, data_final_str = self._get_datas_referencia()

        try:
            # Converte de DD/MM/YYYY para datetime e depois para o formato desejado
            data_inicial = datetime.strptime(data_inicial_str, '%d/%m/%Y').strftime('%d/%m/%Y')
            data_final = datetime.strptime(data_final_str, '%d/%m/%Y').strftime('%d/%m/%Y')
        except Exception:
            # Se já estiverem no formato correto ou houver erro, usa diretamente
            data_inicial, data_final = data_inicial_str, data_final_str

        # Define os caminhos dos arquivos
        base_filename = f"CONCILIACAO_{data_inicial.replace('/', '-')}_a_{data_final.replace('/', '-')}"
        
        if export_type == "fornecedores":
            output_path = self.settings.RESULTS_DIR / f"{base_filename}_FORNECEDORES.xlsx"
        elif export_type == "adiantamentos":
            output_path = self.settings.RESULTS_DIR / f"{base_filename}_ADIANTAMENTOS.xlsx"
        else:
            output_path = self.settings.RESULTS_DIR / f"{base_filename}.xlsx"

        try:
            if not self.conn:
                error_msg = "Tentativa de exportação com conexão fechada"
                logger.error(error_msg)
                raise ResultsSaveError(error_msg, caminho=output_path)
            
            writer = pd.ExcelWriter(output_path, engine='openpyxl')
            
            # Query para obter estatísticas de processamento
            query_stats = f"""
                SELECT 
                    COUNT(*) as total_registros,
                    SUM(CASE WHEN status = 'Conferido' THEN 1 ELSE 0 END) as conciliados_ok,
                    SUM(CASE WHEN status = 'Divergente' THEN 1 ELSE 0 END) as divergentes,
                    SUM(CASE WHEN status = 'Pendente' THEN 1 ELSE 0 END) as pendentes,

                    -- Total Financeiro CORRETO: Soma de (J + K) apenas para NF/FT
                    (
                        SELECT ABS(COALESCE(SUM(COALESCE(tit_vencidos_valor_nominal,0) 
                                                    + COALESCE(titulos_a_vencer_valor_nominal,0)), 0))
                        FROM {self.settings.TABLE_FINANCEIRO}
                        WHERE excluido = 0
                        AND UPPER(tipo_titulo) IN ('NF','FT')
                    ) as total_financeiro,

                    -- Total Contábil CORRETO: Usando CONTAS_ITENS em vez de MODELO1
                    (
                        SELECT COALESCE(SUM(ci.saldo_atual), 0)
                        FROM {self.settings.TABLE_CONTAS_ITENS} ci
                        WHERE ci.conta_contabil LIKE '2.01.02.01.0001%'
                    ) as total_contabil,

                    -- Diferença CORRETA
                    (
                    (SELECT ABS(COALESCE(SUM(COALESCE(tit_vencidos_valor_nominal,0) 
                                                    + COALESCE(titulos_a_vencer_valor_nominal,0)), 0))
                        FROM {self.settings.TABLE_FINANCEIRO}
                        WHERE excluido = 0
                        AND UPPER(tipo_titulo) IN ('NF','FT'))
                        -
                        ABS((SELECT COALESCE(SUM(ci.saldo_atual), 0)
                        FROM {self.settings.TABLE_CONTAS_ITENS} ci
                        WHERE ci.conta_contabil LIKE '2.01.02.01.0001%'))
                    ) as diferenca_geral,

                    -- Total de divergência
                    SUM(CASE WHEN status = 'Divergente' THEN diferenca ELSE 0 END) as total_divergencia

                FROM 
                    {self.settings.TABLE_RESULTADO}
            """

            stats = pd.read_sql(query_stats, self.conn).iloc[0]

            # ABA: "Fornecedores Nacionais" (Dados Financeiros) - APENAS PARA FORNECEDORES
            if export_type in ["all", "fornecedores"]:
                query_financeiro = f"""
                    SELECT 
                        fornecedor as "Fornecedor",
                        titulo as "Título",
                        parcela as "Parcela",
                        tipo_titulo as "Tipo Título",
                        CASE 
                            WHEN data_emissao IS NULL OR data_emissao = '' THEN NULL
                            ELSE data_emissao 
                        END as "Data Emissão",
                        CASE 
                            WHEN data_vencimento IS NULL OR data_vencimento = '' THEN NULL
                            ELSE data_vencimento 
                        END as "Data Vencimento",
                        valor_original as "Valor Original",
                        tit_vencidos_valor_nominal as "Títulos Vencidos",  -- NOVA COLUNA J
                        titulos_a_vencer_valor_nominal as "Títulos a Vencer",  -- NOVA COLUNA K
                        (COALESCE(tit_vencidos_valor_nominal, 0) + COALESCE(titulos_a_vencer_valor_nominal, 0)) as "Saldo Devedor",  -- CALCULADO
                        situacao as "Situação",
                        conta_contabil as "Conta Contábil",
                        centro_custo as "Centro Custo"
                    FROM 
                        {self.settings.TABLE_FINANCEIRO}
                    WHERE 
                        excluido = 0
                        AND UPPER(tipo_titulo) IN ('NF','FT')
                    ORDER BY 
                        fornecedor, titulo, parcela
                """
                df_financeiro = pd.read_sql(query_financeiro, self.conn)
                
                # APLICAR SEPARAÇÃO DE CÓDIGO E DESCRIÇÃO
                df_financeiro = self.separar_codigo_descricao(df_financeiro, "Fornecedor", "Código", "Descrição Fornecedor")
                
                # Reorganizar colunas para ter Código e Descrição primeiro
                colunas_ordenadas = ["Código", "Descrição Fornecedor"] + [col for col in df_financeiro.columns if col not in ["Código", "Descrição Fornecedor", "Fornecedor"]]
                df_financeiro = df_financeiro[colunas_ordenadas]

                # Verifica se há problemas com as datas
                logger.info(f"Total de registros financeiros: {len(df_financeiro)}")
                df_financeiro['Data Emissão'] = pd.to_datetime(df_financeiro['Data Emissão'], errors='coerce').dt.strftime('%d/%m/%Y')
                df_financeiro['Data Vencimento'] = pd.to_datetime(df_financeiro['Data Vencimento'], errors='coerce').dt.strftime('%d/%m/%Y')

                df_financeiro.to_excel(writer, sheet_name='Fornecedores Nacionais', index=False)

            # ABA: "Adiantamento de Fornecedores Nacionais" (Dados Financeiros) - APENAS PARA ADIANTAMENTOS
            if export_type in ["all", "adiantamentos"]:
                query_adi_financeiro = f"""
                    SELECT 
                        fornecedor as "Fornecedor",
                        titulo as "Título",
                        parcela as "Parcela",
                        tipo_titulo as "Tipo Título",
                        CASE 
                            WHEN data_emissao IS NULL OR data_emissao = '' THEN NULL
                            ELSE data_emissao 
                        END as "Data Emissão",
                        CASE 
                            WHEN data_vencimento IS NULL OR data_vencimento = '' THEN NULL
                            ELSE data_vencimento 
                        END as "Data Vencimento",
                        valor_original as "Valor Original",
                        tit_vencidos_valor_nominal as "Títulos Vencidos",
                        titulos_a_vencer_valor_nominal as "Títulos a Vencer",
                        (COALESCE(tit_vencidos_valor_nominal, 0) + COALESCE(titulos_a_vencer_valor_nominal, 0)) as "Saldo Devedor",  -- CALCULADO
                        situacao as "Situação",
                        conta_contabil as "Conta Contábil",
                        centro_custo as "Centro Custo"
                    FROM 
                        {self.settings.TABLE_FINANCEIRO}
                    WHERE 
                        excluido = 0
                        AND UPPER(tipo_titulo) IN ('NDF', 'PA')
                    ORDER BY 
                        fornecedor, titulo, parcela
                """
                df_adi_financeiro = pd.read_sql(query_adi_financeiro, self.conn)
                
                # APLICAR SEPARAÇÃO DE CÓDIGO E DESCRIÇÃO
                df_adi_financeiro = self.separar_codigo_descricao(df_adi_financeiro, "Fornecedor", "Código", "Descrição Fornecedor")
                
                # Reorganizar colunas
                colunas_ordenadas = ["Código", "Descrição Fornecedor"] + [col for col in df_adi_financeiro.columns if col not in ["Código", "Descrição Fornecedor", "Fornecedor"]]
                df_adi_financeiro = df_adi_financeiro[colunas_ordenadas]
                
                # Verifica se há problemas com as datas
                logger.info(f"Total de registros financeiros de adiantamento: {len(df_adi_financeiro)}")
                df_adi_financeiro['Data Emissão'] = pd.to_datetime(df_adi_financeiro['Data Emissão'], errors='coerce').dt.strftime('%d/%m/%Y')
                df_adi_financeiro['Data Vencimento'] = pd.to_datetime(df_adi_financeiro['Data Vencimento'], errors='coerce').dt.strftime('%d/%m/%Y')

                df_adi_financeiro.to_excel(writer, sheet_name='Adiantamento de Fornecedores Nacionais', index=False)

            # ABA: "Balancete" (Dados Contábeis) - APENAS PARA FORNECEDORES
            if export_type in ["all", "fornecedores"]:
                query_contabil = f"""
                    SELECT 
                        conta_contabil as "Conta Contábil",
                        descricao_conta as "Descrição",
                        codigo_fornecedor as "Código Fornecedor", 
                        descricao_fornecedor as "Descrição Fornecedor",
                        saldo_anterior as "Saldo Anterior",
                        debito as "Débito",
                        credito as "Crédito", 
                        saldo_atual as "Saldo Atual",
                        tipo_fornecedor as "Tipo"
                    FROM 
                        {self.settings.TABLE_MODELO1}
                    WHERE 
                        (descricao_conta LIKE '%FORNEC%' OR tipo_fornecedor LIKE '%FORNEC%')
                        AND conta_contabil LIKE '2.01.02.01.0001%'
                        AND (debito != 0 OR credito != 0 OR saldo_atual != 0)  -- Só mostra movimentos com valores
                    ORDER BY 
                        ABS(saldo_atual) DESC,  -- Maiores saldos primeiro
                        conta_contabil, 
                        codigo_fornecedor
                """
                df_contabil = pd.read_sql(query_contabil, self.conn)

                # Remove a coluna 'ordem' apenas se ela existir
                if 'ordem' in df_contabil.columns:
                    df_contabil.drop(columns=["ordem"], inplace=True)

                # Aplica formatação monetária nas colunas numéricas
                monetary_columns = ['Saldo anterior', 'Debito', 'Credito', 'Saldo atual']
                for col in monetary_columns:
                    if col in df_contabil.columns:
                        df_contabil[col] = df_contabil[col].apply(self.formatar_credito)

                df_contabil.to_excel(writer, sheet_name='Balancete', index=False)
                
            # ABA: "Adiantamento" (Dados de Adiantamentos) - APENAS PARA ADIANTAMENTOS
            if export_type in ["all", "adiantamentos"]:
                query_adiantamento = f"""
                    SELECT 
                        conta_contabil as "Conta Contábil",
                        descricao_item as "Descrição Item",
                        codigo_fornecedor as "Código Fornecedor",
                        descricao_fornecedor as "Descrição Fornecedor",
                        saldo_anterior as "Saldo Anterior",
                        saldo_atual as "Saldo Atual"
                    FROM 
                        {self.settings.TABLE_ADIANTAMENTO}
                    WHERE 
                        -- FILTRO: Remove linhas vazias/inválidas (MESMO FILTRO DA CONTAS X ITENS)
                        (descricao_fornecedor IS NOT NULL AND descricao_fornecedor != '')
                        AND (saldo_anterior IS NOT NULL AND saldo_anterior != 0)
                        AND (saldo_atual IS NOT NULL AND saldo_atual != 0)
                    ORDER BY 
                        conta_contabil, codigo_fornecedor
                """
                df_adiantamento = pd.read_sql(query_adiantamento, self.conn)
                
                # APLICAR SEPARAÇÃO SE NECESSÁRIO
                if "Código Fornecedor" in df_adiantamento.columns:
                    df_adiantamento = self.separar_codigo_descricao(df_adiantamento, "Código Fornecedor", "Código", "Descrição Fornecedor")
                    
                    # Reorganizar colunas se a separação foi aplicada
                    if "Código" in df_adiantamento.columns and "Descrição Fornecedor" in df_adiantamento.columns:
                        colunas_ordenadas = ["Conta Contábil", "Descrição Item", "Código", "Descrição Fornecedor"] + \
                                        [col for col in df_adiantamento.columns if col not in ["Conta Contábil", "Descrição Item", "Código", "Descrição Fornecedor", "Código Fornecedor"]]
                        df_adiantamento = df_adiantamento[colunas_ordenadas]
                
                df_adiantamento.to_excel(writer, sheet_name='Adiantamento', index=False)
                
            # ABA: "Contas x Itens" (Detalhamento Contábil) - APENAS PARA FORNECEDORES
            if export_type in ["all", "fornecedores"]:
                query_contas_itens = f"""
                    SELECT 
                        conta_contabil as "Conta Contábil",
                        descricao_item as "Descrição Item",
                        codigo_fornecedor as "Código Fornecedor",
                        descricao_fornecedor as "Descrição Fornecedor",
                        saldo_anterior as "Saldo Anterior",
                        saldo_atual as "Saldo Atual"
                    FROM 
                        {self.settings.TABLE_CONTAS_ITENS}
                    WHERE 
                        -- FILTRO: Remove linhas vazias/inválidas
                        (descricao_fornecedor IS NOT NULL AND descricao_fornecedor != '')
                        AND (saldo_anterior IS NOT NULL AND saldo_anterior != 0)
                        AND (saldo_atual IS NOT NULL AND saldo_atual != 0)
                    ORDER BY 
                        conta_contabil, codigo_fornecedor
                """
                df_contas_itens = pd.read_sql(query_contas_itens, self.conn)
                
                # APLICAR SEPARAÇÃO SE NECESSÁRIO
                if "Código Fornecedor" in df_contas_itens.columns:
                    df_contas_itens = self.separar_codigo_descricao(df_contas_itens, "Código Fornecedor", "Código", "Descrição Fornecedor")
                    
                    # Reorganizar colunas se a separação foi aplicada
                    if "Código" in df_contas_itens.columns and "Descrição Fornecedor" in df_contas_itens.columns:
                        colunas_ordenadas = ["Conta Contábil", "Descrição Item", "Código", "Descrição Fornecedor"] + \
                                        [col for col in df_contas_itens.columns if col not in ["Conta Contábil", "Descrição Item", "Código", "Descrição Fornecedor", "Código Fornecedor"]]
                        df_contas_itens = df_contas_itens[colunas_ordenadas]
                
                df_contas_itens.to_excel(writer, sheet_name='Contas x Itens', index=False)

            # ABA: "Resumo Adiantamentos" - APENAS PARA ADIANTAMENTOS
            if export_type in ["all", "adiantamentos"]:
                query_resumo_adiantamento = f"""
                    SELECT 
                        codigo_fornecedor as "Código Fornecedor",
                        descricao_fornecedor as "Descrição Fornecedor",
                        total_financeiro as "Total Financeiro",
                        total_contabil as "Total Contábil",
                        diferenca as "Diferença",
                        status as "Status",
                        detalhes as "Detalhes"
                    FROM 
                        {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
                    WHERE 
                        -- FILTRO DIRETO NO SQL: Remove registros com código vazio
                        codigo_fornecedor IS NOT NULL 
                        AND codigo_fornecedor != ''
                        AND TRIM(codigo_fornecedor) != ''
                    ORDER BY 
                        ABS(diferenca) DESC,
                        codigo_fornecedor
                """
                df_resumo_adiantamento = pd.read_sql(query_resumo_adiantamento, self.conn)
                
                # MESMO TRATAMENTO ROBUSTO PARA ADIANTAMENTOS
                antes = len(df_resumo_adiantamento)
                df_resumo_adiantamento = df_resumo_adiantamento[
                    df_resumo_adiantamento["Código Fornecedor"].notna() & 
                    (df_resumo_adiantamento["Código Fornecedor"].astype(str).str.strip() != "") &
                    (df_resumo_adiantamento["Código Fornecedor"].astype(str).str.strip() != "None") &
                    (df_resumo_adiantamento["Código Fornecedor"].astype(str).str.strip() != "nan")
                ]
                depois = len(df_resumo_adiantamento)
                
                if antes > depois:
                    logger.warning(f"Removidas {antes - depois} linhas com código vazio do resumo de adiantamentos")
                
                # APLICAR SEPARAÇÃO SE A COLUNA CÓDIGO FORNECEDOR CONTÉM CÓDIGO-DESCRIÇÃO
                if "Código Fornecedor" in df_resumo_adiantamento.columns and len(df_resumo_adiantamento) > 0:
                    df_resumo_adiantamento = self.separar_codigo_descricao(df_resumo_adiantamento, "Código Fornecedor", "Código", "Descrição Fornecedor")
                    
                    # FILTRO EXTRA APÓS SEPARAÇÃO
                    antes_sep = len(df_resumo_adiantamento)
                    df_resumo_adiantamento = df_resumo_adiantamento[
                        df_resumo_adiantamento["Código"].notna() & 
                        (df_resumo_adiantamento["Código"].astype(str).str.strip() != "") &
                        (df_resumo_adiantamento["Código"].astype(str).str.strip() != "None") &
                        (df_resumo_adiantamento["Código"].astype(str).str.strip() != "nan")
                    ]
                    depois_sep = len(df_resumo_adiantamento)
                    
                    if antes_sep > depois_sep:
                        logger.warning(f"Removidas {antes_sep - depois_sep} linhas com código vazio após separação (adiantamentos)")
                    
                    # Reorganizar colunas
                    colunas_ordenadas = ["Código", "Descrição Fornecedor"] + [col for col in df_resumo_adiantamento.columns if col not in ["Código", "Descrição Fornecedor", "Código Fornecedor"]]
                    df_resumo_adiantamento = df_resumo_adiantamento[colunas_ordenadas]
                
                if len(df_resumo_adiantamento) > 0:
                    df_resumo_adiantamento.to_excel(writer, sheet_name='Resumo Adiantamentos', index=False)
                else:
                    colunas = ["Código", "Descrição Fornecedor", "Total Financeiro", "Total Contábil", "Diferença", "Status", "Detalhes"]
                    df_resumo_adiantamento = pd.DataFrame(columns=colunas)
                    df_resumo_adiantamento.to_excel(writer, sheet_name='Resumo Adiantamentos', index=False)
                    logger.warning("Resumo Adiantamentos está vazio - criando planilha vazia")

            # ABA: "Resumo da Conciliação" (Principal) - APENAS PARA FORNECEDORES
            if export_type in ["all", "fornecedores"]:
                query_resumo = f"""
                    SELECT 
                        codigo_fornecedor as "Código Fornecedor",
                        descricao_fornecedor as "Descrição Fornecedor",
                        saldo_financeiro as "Total Financeiro",
                        saldo_contabil as "Total Contábil",
                        diferenca as "Diferença",
                        status as "Status",
                        detalhes as "Detalhes"
                    FROM 
                        {self.settings.TABLE_RESULTADO}
                    WHERE 
                        -- FILTRO DIRETO NO SQL: Remove registros com código vazio
                        codigo_fornecedor IS NOT NULL 
                        AND codigo_fornecedor != ''
                        AND TRIM(codigo_fornecedor) != ''
                    ORDER BY 
                        ABS(diferenca) DESC,
                        codigo_fornecedor
                """
                df_resumo = pd.read_sql(query_resumo, self.conn)
                
                # LOG DETALHADO PARA DEBUG
                logger.info(f"Total de registros no resumo após filtro SQL: {len(df_resumo)}")
                
                # VERIFICAÇÃO EXTRA: Remove qualquer linha que ainda possa ter código vazio
                antes = len(df_resumo)
                df_resumo = df_resumo[
                    df_resumo["Código Fornecedor"].notna() & 
                    (df_resumo["Código Fornecedor"].astype(str).str.strip() != "") &
                    (df_resumo["Código Fornecedor"].astype(str).str.strip() != "None") &
                    (df_resumo["Código Fornecedor"].astype(str).str.strip() != "nan")
                ]
                depois = len(df_resumo)
                
                if antes > depois:
                    logger.warning(f"Removidas {antes - depois} linhas com código vazio após filtro extra")
                
                # LOG DOS PRIMEIROS REGISTROS PARA VERIFICAÇÃO
                if len(df_resumo) > 0:
                    logger.info(f"Primeiros 5 códigos no resumo: {df_resumo['Código Fornecedor'].head().tolist()}")
                
                # APLICAR SEPARAÇÃO SE A COLUNA CÓDIGO FORNECEDOR CONTÉM CÓDIGO-DESCRIÇÃO
                if "Código Fornecedor" in df_resumo.columns and len(df_resumo) > 0:
                    df_resumo = self.separar_codigo_descricao(df_resumo, "Código Fornecedor", "Código", "Descrição Fornecedor")
                    
                    # FILTRO EXTRA APÓS SEPARAÇÃO: Remove linhas onde o código ficou vazio após separação
                    antes_sep = len(df_resumo)
                    df_resumo = df_resumo[
                        df_resumo["Código"].notna() & 
                        (df_resumo["Código"].astype(str).str.strip() != "") &
                        (df_resumo["Código"].astype(str).str.strip() != "None") &
                        (df_resumo["Código"].astype(str).str.strip() != "nan")
                    ]
                    depois_sep = len(df_resumo)
                    
                    if antes_sep > depois_sep:
                        logger.warning(f"Removidas {antes_sep - depois_sep} linhas com código vazio após separação")
                    
                    # Reorganizar colunas
                    colunas_ordenadas = ["Código", "Descrição Fornecedor"] + [col for col in df_resumo.columns if col not in ["Código", "Descrição Fornecedor", "Código Fornecedor"]]
                    df_resumo = df_resumo[colunas_ordenadas]

                # Garantir que as colunas sejam float antes de exportar
                for col in ["Total Financeiro", "Total Contábil", "Diferença"]:
                    if col in df_resumo.columns:
                        df_resumo[col] = pd.to_numeric(df_resumo[col], errors="coerce").fillna(0)

                # LOG FINAL ANTES DE EXPORTAR
                logger.info(f"Total final de registros no resumo: {len(df_resumo)}")
                
                if len(df_resumo) > 0:
                    df_resumo.to_excel(writer, sheet_name='Resumo da Conciliação', index=False)
                else:
                    # Cria um DataFrame vazio com as colunas corretas para evitar erro
                    colunas = ["Código", "Descrição Fornecedor", "Total Financeiro", "Total Contábil", "Diferença", "Status", "Detalhes"]
                    df_resumo = pd.DataFrame(columns=colunas)
                    df_resumo.to_excel(writer, sheet_name='Resumo da Conciliação', index=False)
                    logger.warning("Resumo da Conciliação está vazio - criando planilha vazia")

            # Query para estatísticas de adiantamentos
            query_adiantamento_stats = f"""
                SELECT 
                    COUNT(*) as total_adiantamentos,
                    SUM(CASE WHEN status = 'Conferido' THEN 1 ELSE 0 END) as adiantamentos_ok,
                    SUM(CASE WHEN status = 'Divergente' THEN 1 ELSE 0 END) as adiantamentos_divergentes,
                    SUM(CASE WHEN status = 'Pendente' THEN 1 ELSE 0 END) as adiantamentos_pendentes,
                    ABS(SUM(total_financeiro)) as total_financeiro_adiantamento,
                    SUM(total_contabil) as total_contabil_adiantamento,
                    (
                        ABS(SUM(total_financeiro)) - SUM(total_contabil)
                    ) as diferenca_adiantamento
                FROM 
                    {self.settings.TABLE_RESULTADO_ADIANTAMENTO}
            """

            adiantamento_stats = pd.read_sql(query_adiantamento_stats, self.conn).iloc[0]

            # Cria aba de Metadados específica para cada tipo
            if export_type == "fornecedores":
                metadata_items = [
                    'Data e Hora do Processamento',
                    'Período de Referência',
                    'Total de Fornecedores Processados',
                    'Conciliações Conferidas',
                    'Conciliações Divergentes',
                    'Conciliações Pendentes',
                    'Total Financeiro (R$)', 
                    'Total Contábil (R$)',
                    'Diferença Total (R$)',
                    '--- CONFIGURAÇÕES ---',
                    'Legenda de Status',
                    'Tolerância de Diferença'
                ]

                metadata_values = [
                    datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                    f'{data_inicial} a {data_final}',
                    int(stats['total_registros']),
                    int(stats['conciliados_ok']),
                    int(stats['divergentes']),
                    int(stats['pendentes']),
                    f"R$ {stats['total_financeiro']:,.2f}",
                    f"R$ {stats['total_contabil']:,.2f}",
                    f"R$ {stats['diferenca_geral']:,.2f}",
                    '---',
                    'CONFERIDO: Diferença dentro da tolerância (até 3%) | DIVERGENTE: Diferença significativa | PENDENTE: Sem correspondência',
                    'Até 3% de discrepância é considerada tolerável'
                ]
                
            elif export_type == "adiantamentos":
                metadata_items = [
                    'Data e Hora do Processamento',
                    'Período de Referência',
                    'Total de Adiantamentos Processados',
                    'Adiantamentos Conferidos',
                    'Adiantamentos Divergentes',
                    'Adiantamentos Pendentes',
                    'Total Financeiro Adiantamentos (R$)',
                    'Total Contábil Adiantamentos (R$)',
                    'Diferença Total Adiantamentos (R$)',
                    '--- CONFIGURAÇÕES ---',
                    'Legenda de Status',
                    'Tolerância de Diferença'
                ]

                metadata_values = [
                    datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                    f'{data_inicial} a {data_final}',
                    int(adiantamento_stats['total_adiantamentos']),
                    int(adiantamento_stats['adiantamentos_ok']),
                    int(adiantamento_stats['adiantamentos_divergentes']),
                    int(adiantamento_stats['adiantamentos_pendentes']),
                    f"R$ {adiantamento_stats['total_financeiro_adiantamento']:,.2f}",
                    f"R$ {adiantamento_stats['total_contabil_adiantamento']:,.2f}",
                    f"R$ {adiantamento_stats['diferenca_adiantamento']:,.2f}",
                    '---',
                    'CONFERIDO: Diferença dentro da tolerância (até 3%) | DIVERGENTE: Diferença significativa | PENDENTE: Sem correspondência',
                    'Até 3% de discrepância é considerada tolerável'
                ]
                
            else:  # all
                metadata_items = [
                    'Data e Hora do Processamento',
                    'Período de Referência',
                    'Total de Fornecedores Processados',
                    'Conciliações Conferidas',
                    'Conciliações Divergentes',
                    'Conciliações Pendentes',
                    'Total Financeiro (R$)',
                    'Total Contábil (R$)',
                    'Diferença Total (R$)',
                    '--- ADIANTAMENTOS ---',
                    'Total de Adiantamentos Processados',
                    'Adiantamentos Divergentes', 
                    'Total Financeiro Adiantamentos (R$)',
                    'Total Contábil Adiantamentos (R$)',
                    'Saldo Líquido Adiantamentos (R$)',
                    '--- CONFIGURAÇÕES ---',
                    'Legenda de Status',
                    'Tolerância de Diferença'
                ]

                metadata_values = [
                    datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                    f'{data_inicial} a {data_final}',
                    int(stats['total_registros']),
                    int(stats['conciliados_ok']),
                    int(stats['divergentes']),
                    int(stats['pendentes']),
                    f"R$ {stats['total_financeiro']:,.2f}",
                    f"R$ {stats['total_contabil']:,.2f}",
                    f"R$ {stats['diferenca_geral']:,.2f}",
                    '---',
                    int(adiantamento_stats['total_adiantamentos']),
                    int(adiantamento_stats['adiantamentos_divergentes']),
                    f"R$ {adiantamento_stats['total_financeiro_adiantamento']:,.2f}",
                    f"R$ {adiantamento_stats['total_contabil_adiantamento']:,.2f}",
                    f"R$ {adiantamento_stats['diferenca_adiantamento']:,.2f}",
                    '---',
                    'CONFERIDO: Diferença dentro da tolerância (até 3%) | DIVERGENTE: Diferença significativa | PENDENTE: Sem correspondência',
                    'Até 3% de discrepância é considerada tolerável'
                ]

            # VERIFICAÇÃO DE COMPRIMENTO
            if len(metadata_items) != len(metadata_values):
                logger.error(f"Metadados incompatíveis: {len(metadata_items)} itens vs {len(metadata_values)} valores")
                min_length = min(len(metadata_items), len(metadata_values))
                metadata_items = metadata_items[:min_length]
                metadata_values = metadata_values[:min_length]
                
            metadata = {
                'Item': metadata_items,
                'Valor': metadata_values
            }

            df_metadata = pd.DataFrame(metadata)
            df_metadata.to_excel(writer, sheet_name='Metadados', index=False)
            writer.close()
            
            # Aplica estilos ao arquivo gerado
            workbook = openpyxl.load_workbook(output_path)
            
            # Aplica estilos à aba Metadados
            if "Metadados" in workbook.sheetnames:
                meta_sheet = workbook["Metadados"]
                for row_idx, (item, value) in enumerate(zip(metadata_items, metadata_values), 1):
                    meta_sheet.cell(row=row_idx, column=1, value=item)
                    meta_sheet.cell(row=row_idx, column=2, value=value)
                self._apply_metadata_styles(meta_sheet, metadata_items, metadata_values)
            
            # Aplica estilos melhorados às abas principais
            if export_type in ["all", "fornecedores"] and 'Resumo da Conciliação' in workbook.sheetnames:
                resumo_sheet = workbook['Resumo da Conciliação']
                self._apply_enhanced_styles(resumo_sheet, stats)
                resumo_sheet.auto_filter.ref = resumo_sheet.dimensions
                
            if export_type in ["all", "adiantamentos"] and 'Resumo Adiantamentos' in workbook.sheetnames:
                adiantamento_sheet = workbook['Resumo Adiantamentos']
                self._apply_enhanced_styles(adiantamento_sheet, adiantamento_stats)
                adiantamento_sheet.auto_filter.ref = adiantamento_sheet.dimensions
            
            # Aplica estilos básicos às outras abas
            for sheetname in workbook.sheetnames:
                if sheetname not in ['Resumo da Conciliação', 'Resumo Adiantamentos', 'Metadados']:
                    sheet = workbook[sheetname]
                    self._apply_styles(sheet)
            
            # Protege todas as abas
            self._protect_sheets(workbook)
            
            workbook.save(output_path)
            
            # Valida o arquivo gerado
            if not self.validate_output(output_path, export_type):
                raise ValueError("A validação da planilha gerada falhou")
                        
            logger.info(f"Arquivo exportado com sucesso: {output_path}")
            return output_path
            
        except Exception as e:
            error_msg = f"Erro ao exportar resultados: {e}"
            logger.error(error_msg)
            raise ResultsSaveError(error_msg, caminho=output_path) from e
    

    def validate_output(self, output_path, export_type="all"):
        """
        Valida a estrutura do arquivo Excel gerado.
        
        Args:
            output_path: Caminho do arquivo a ser validado
            export_type: Tipo de exportação (opcional, padrão "all")
        """
        try:
            wb = openpyxl.load_workbook(output_path)
            
            # Define as abas obrigatórias baseadas no tipo de exportação
            if export_type == "fornecedores":
                required_sheets = ['Resumo da Conciliação', 'Fornecedores Nacionais', 'Balancete', 'Contas x Itens', 'Metadados']
            elif export_type == "adiantamentos":
                required_sheets = ['Resumo Adiantamentos', 'Adiantamento de Fornecedores Nacionais', 'Adiantamento', 'Metadados']
            else:
                required_sheets = ['Resumo da Conciliação', 'Fornecedores Nacionais', 'Balancete', 'Contas x Itens', 
                                'Resumo Adiantamentos', 'Adiantamento de Fornecedores Nacionais', 'Adiantamento', 'Metadados']
            
            for sheet in required_sheets:
                if sheet not in wb.sheetnames:
                    raise ValueError(f"Aba '{sheet}' não encontrada no arquivo gerado")
            
            # Verifica formatação monetária nas abas principais
            if export_type in ["all", "fornecedores"] and 'Resumo da Conciliação' in wb.sheetnames:
                resumo = wb['Resumo da Conciliação']
                monetary_columns = ['Saldo Financeiro', 'Saldo Contábil', 'Diferença']
                
                header = [cell.value for cell in resumo[1] if cell.value is not None]
                
                for col_name in monetary_columns:
                    if col_name in header:
                        col_idx = header.index(col_name) + 1
                        sample_cell = resumo.cell(row=2, column=col_idx)
                        if sample_cell.value is not None and hasattr(sample_cell, 'number_format'):
                            if 'R$' not in sample_cell.number_format and '#,##0.00' not in sample_cell.number_format:
                                logger.warning(f"Coluna '{col_name}' não está formatada como moeda brasileira")
            
            if export_type in ["all", "adiantamentos"] and 'Resumo Adiantamentos' in wb.sheetnames:
                adiantamento = wb['Resumo Adiantamentos']
                monetary_columns = ['Total Financeiro', 'Total Contábil', 'Diferença']
                
                header = [cell.value for cell in adiantamento[1] if cell.value is not None]
                
                for col_name in monetary_columns:
                    if col_name in header:
                        col_idx = header.index(col_name) + 1
                        sample_cell = adiantamento.cell(row=2, column=col_idx)
                        if sample_cell.value is not None and hasattr(sample_cell, 'number_format'):
                            if 'R$' not in sample_cell.number_format and '#,##0.00' not in sample_cell.number_format:
                                logger.warning(f"Coluna '{col_name}' não está formatada como moeda brasileira")
            
            return True
            
        except Exception as e:
            error_msg = f"Validação falhou: {e}"
            logger.error(error_msg)
            return False
    def close(self):
        """Fecha a conexão com o banco de dados"""
        if self.conn:
            try:
                self.conn.close()
                logger.info("Conexão com o banco de dados fechada")
            except Exception as e:
                logger.error(f"Erro ao fechar conexão: {e}")

    def __enter__(self):
        """
        Suporte para context manager (with statement).
        
        Returns:
            DatabaseManager: Instância da classe
        """
        self._initialize_database()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """
        Suporte para context manager (with statement).
        
        Args:
            exc_type: Tipo de exceção (se ocorreu)
            exc_val: Valor da exceção
            exc_tb: Traceback da exceção
        """
        self.close()